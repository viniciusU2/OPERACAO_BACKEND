from typing import List

from fastapi import APIRouter, Depends, HTTPException
from pymysql import IntegrityError
from sqlalchemy import text
from sqlalchemy.orm import Session
from database import get_db
from models.OS_models import OrdemServico
from OS.schemas import (
    BaixaOSLoteResponse,
    BaixaOSLoteTipoAtivo,
    OrdemServicoCreate,
    OrdemServicoCreateLote,
    OrdemServicoResponse,
    OrdemServicoUpdate,
)
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from database import get_db
from models.Ativo import Ativo
from models.instalacao_models import Subestacao
from models.SS_models import SolicitacaoServico
from models import OS_models
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import shutil
from datetime import datetime
import os
import re
from fastapi.background import BackgroundTasks
from typing import List
from fastapi.responses import FileResponse
from auth.dependencies import require_roles
from utils.documentos_operacao import (
    especie_documento_por_ativo,
    normalizar_prioridade_operacao,
)



router = APIRouter(prefix="/os", tags=["Ordem de Serviço"])

def remover_arquivo(path: str):
    if os.path.exists(path):
        os.remove(path)


def nome_arquivo_seguro(texto: str):
    if not texto:
        return "sem_nome"
    
    # substitui tudo que não for letra/número por _
    return re.sub(r'[^A-Za-z0-9_-]', '_', texto)

MAPEAMENTO_CELULAS = {
    "NUM_OS": "H1",
    "NUM_SI": "H2",
    "ESPECIE": "A4",
    "CODIGO_ATIVO": "D4",
    "NUM_APR": "G4",
    "INSTALACAO": "A6",
    "LOCALIZACAO": "D6",
    "COMPLEMENTO": "G6",
    "ORIGENS": "A8",
    "DEFEITO": "F8",
    "ESQUEMA_SERVICOS": "A10",
    "PRIORIDADES": "F10",
    "RESPONSAVEL": "A12",
    "RESPONSAVEL_MANUTENCAO": "D26",
    "RESPONSAVEL_OPERACAO": "D32",
    "SUBSTITUTO": "F12",
    "DT_INICIO_PROGRAMADO": "A14",
    "DT_FIM_PROGRAMADO": "F14",
    "DESC_SERVICOS": "A16",
    "OBSERVACOES": "A18",
    "CAUSAS": "A20",
    "CAUSAS_SEC": "F20",
    "DT_ABERTURA_SS": "A22",
    "DT_INICIO_EXEC": "D22",
    "DT_FIM_EXEC": "G22",
    "DT_INICIO_EXEC2": "D27",
    "DT_FIM_EXEC2": "D33",
    "CENTRO_CUSTOS": "A24",
   
}


def limpar(valor):
    if valor is None:

        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    return str(valor)

def gerar_xlsm(modelo, destino, contexto, mapeamento):
    shutil.copy(modelo, destino)
    img = Image("modelos/logo.jpg")  # caminho da sua logo
    # 🔽 Redimensiona a imagem
    img.width = 150
    img.height = 45

    wb = load_workbook(destino, keep_vba=True)
    ws = wb.active
    ws.add_image(img, "A1")
    
    if contexto.get("STATUS", "") == "ENCERRADA":
     ws["H26"] = "INICIADA"
     ws["H32"] = "ENCERRADA"

    for campo, celula in mapeamento.items():
        ws[celula] = contexto.get(campo, "")

    wb.save(destino)

def montar_contexto_os(os, ativo=None):

    codigo_ativo = None

    if ativo:
        codigo_ativo = ativo.codigo_ativo
        local =  ativo.vao
        fase = ativo.fase
    return {
        # ================= IDENTIFICAÇÃO =================
        "NUM_OS": limpar(os.numero_os),
        "NUM_SI": limpar(os.numero_si),
        "ESPECIE": limpar(os.especie),

        # 🔥 AQUI ESTÁ A MÁGICA
        "CODIGO_ATIVO": limpar(codigo_ativo),

        "NUM_APR": limpar(os.numero_apr),

        # ================= LOCALIZAÇÃO =================
        "INSTALACAO": limpar(os.instalacao),
        "LOCALIZACAO": limpar(local),
        "COMPLEMENTO": limpar(fase),

        # ================= DESCRIÇÃO =================
        "ORIGENS": limpar(os.origens),
        "DEFEITO": limpar(os.defeito),
        "ESQUEMA_SERVICOS": limpar(os.esquema_servicos),
        "DESC_SERVICOS": limpar(os.descricao_servicos),
        "OBSERVACOES": limpar(os.observacoes),

        # ================= CAUSAS =================
        "CAUSAS": limpar(os.causa_primaria),
        "CAUSAS_SEC": limpar(os.causa_secundaria),

        # ================= CONTROLE =================
        "PRIORIDADES": limpar(os.prioridade),
        "RESPONSAVEL": limpar(os.responsavel),
        "SUBSTITUTO": limpar(os.substituto),
        "RESPONSAVEL_MANUTENCAO": limpar(os.responsavel_manutencao),
        "RESPONSAVEL_OPERACAO": limpar(os.responsavel_operacao),
        "CENTRO_CUSTOS": limpar(os.centro_custos),
      

        # ================= DATAS =================
        "DT_ABERTURA_SS": limpar(os.data_abertura_ss),
        "DT_INICIO_PROGRAMADO": limpar(os.data_inicio_programado),
        "DT_FIM_PROGRAMADO": limpar(os.data_fim_programado),
        "DT_INICIO_EXEC": limpar(os.data_inicio_execucao),
        "DT_FIM_EXEC": limpar(os.data_fim_execucao),
        "DT_INICIO_EXEC2": limpar(os.data_inicio_execucao),
        "DT_FIM_EXEC2": limpar(os.data_fim_execucao),
        "STATUS": limpar(os.status),  # ✅ AQUI
    }

# 🔹 Lista fixa de subestações
SUBESTACOES_SIGLAS = ["BJD", "GOR", "JAB"]


# =========================
# 🔥 GERAR NUMERO OS
# =========================
def gerar_numero_os(db: Session, sigla: str, codigo_ativo: str | None) -> tuple[str, str]:
    ano_atual = datetime.now().year

    # 🔹 Buscar OS existentes da mesma subestação e ano
    registros = (
        db.query(OS_models.OrdemServico.numero_os)
        .filter(OS_models.OrdemServico.numero_os.like(f"OS-{sigla}-%-{ano_atual}%"))
        .all()
    )

    numeros = []

    for (numero_os,) in registros:
        match = re.search(rf"OS-{sigla}-(\d+)-{ano_atual}", numero_os)
        if match:
            numeros.append(int(match.group(1)))

    proximo = max(numeros) + 1 if numeros else 1
    numero_formatado = str(proximo).zfill(4)

    # 🔹 Sanitizar código do ativo (opcional mas recomendado)
    numero_os = f"OS-{sigla}-{numero_formatado}-{ano_atual}"
    numero_apr = f"APR-{sigla}-{numero_formatado}-{ano_atual}"

    if codigo_ativo:
        codigo_ativo = re.sub(r"[^A-Za-z0-9\-]", "", codigo_ativo)
        numero_os = f"{numero_os}-{codigo_ativo}"

    return numero_os, numero_apr




# =========================
# 🔥 CRIAR ORDEM DE SERVIÇO
# =========================
@router.post("", response_model=OrdemServicoCreate)
def criar_ordem_servico(
    os_data: OrdemServicoCreate,
    db: Session = Depends(get_db)
):
    print(os_data)

    # 🔹 Validação de datas
    if (
        os_data.data_inicio_programado
        and os_data.data_fim_programado
        and os_data.data_fim_programado < os_data.data_inicio_programado
    ):
        raise HTTPException(
            status_code=400,
            detail="Data final não pode ser anterior à data inicial"
        )

    # 🔹 Validar Subestação
    if not os_data.id_subestacao:
        raise HTTPException(400, "Subestação é obrigatória")

    sub = db.query(Subestacao).filter(
        Subestacao.id_subestacao == os_data.id_subestacao
    ).first()

    if not sub:
        raise HTTPException(400, "Subestação inválida")

    # 🔹 Define nome da instalação
    os_data.instalacao = sub.nome

    # 🔹 Definir sigla (baseado na lista)
    try:
        sigla = SUBESTACOES_SIGLAS[os_data.id_subestacao - 1]
    except IndexError:
        raise HTTPException(400, "Subestação sem sigla configurada")

    # 🔹 Validar Ativo
    ativo = None
    codigo_ativo = None

    if os_data.id_ativo:
        ativo = db.query(Ativo).filter(
            Ativo.id_ativo == os_data.id_ativo
        ).first()

        if not ativo:
            raise HTTPException(400, "Ativo inválido")

        codigo_ativo = ativo.codigo_ativo

    # 🔥 GERAR NUMERO OS
    if ativo:
        os_data.especie = especie_documento_por_ativo(ativo) or os_data.especie

    os_data.prioridade = normalizar_prioridade_operacao(os_data.prioridade)
    os_data.numero_os,  os_data.numero_apr = gerar_numero_os(db, sigla, codigo_ativo)

    # 🔹 Criar OS
    data = os_data.dict(exclude={"codigo_ativo"})
    nova_os = OS_models.OrdemServico(**data)
    print(data)

    db.add(nova_os)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Erro ao gerar número da OS. Tente novamente."
        )

    db.refresh(nova_os)

    # -------------------------
    # 🔥 GERAR XLSM
    # -------------------------
    pasta_saida = "saida"
    os.makedirs(pasta_saida, exist_ok=True)

    contexto = montar_contexto_os(os_data, ativo)

    numero_os_safe = nome_arquivo_seguro(nova_os.numero_os)
    nome_arquivo = f"{numero_os_safe}.xlsm"
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    gerar_xlsm(
        modelo="modelos/MODELO_OS.xlsx",
        destino=caminho_saida,
        contexto=contexto,
        mapeamento=MAPEAMENTO_CELULAS
    )

    return nova_os



@router.get("/{id_os}/download")
def baixar_os(id_os: int, db: Session = Depends(get_db)):

    # 🔹 Buscar OS
    os_db = (
        db.query(OS_models.OrdemServico)
        .filter(OS_models.OrdemServico.id_os == id_os)
        .first()
    )

    if not os_db:
        raise HTTPException(
            status_code=404,
            detail="OS não encontrada"
        )

    # 🔹 Buscar ativo (se existir)
    ativo = None
    if os_db.id_ativo:
        ativo = db.query(Ativo).filter(
            Ativo.id_ativo == os_db.id_ativo
        ).first()

    # 🔹 Montar contexto
    contexto = montar_contexto_os(os_db, ativo)

    # 🔹 Nome seguro
    numero_os_safe = nome_arquivo_seguro(os_db.numero_os)
    nome_arquivo = f"{numero_os_safe}.xlsm"

    # 🔹 Caminho temporário
    pasta_temp = "temp"
    os.makedirs(pasta_temp, exist_ok=True)

    caminho_saida = os.path.join(pasta_temp, nome_arquivo)

    # 🔹 Gerar arquivo na hora
    gerar_xlsm(
        modelo="modelos/MODELO_OS.xlsx",  # ⚠️ importante ser XLSM
        destino=caminho_saida,
        contexto=contexto,
        mapeamento=MAPEAMENTO_CELULAS
    )

    # 🔹 Retornar arquivo
    return FileResponse(
        path=caminho_saida,
        filename=nome_arquivo,
        media_type="application/octet-stream"
    )


@router.get("/subestacao/1/gerar-os")
def gerar_os_subestacao(db: Session = Depends(get_db)):

    # 🔹 Buscar todas as OS da subestação 1
    ordens = (
        db.query(OS_models.OrdemServico)
        .filter(OS_models.OrdemServico.id_subestacao == 1)
        .all()
    )

    if not ordens:
        raise HTTPException(
            status_code=404,
            detail="Nenhuma OS encontrada para essa subestação"
        )

    # 🔹 Criar pasta de saída
    pasta_saida = "saida"
    os.makedirs(pasta_saida, exist_ok=True)

    arquivos_gerados = []

    for os_db in ordens:

        # 🔹 Buscar ativo (se existir)
        ativo = None
        if os_db.id_ativo:
            ativo = db.query(Ativo).filter(
                Ativo.id_ativo == os_db.id_ativo
            ).first()

        # 🔹 Montar contexto
        contexto = montar_contexto_os(os_db, ativo)

        # 🔹 Nome seguro
        numero_os_safe = nome_arquivo_seguro(os_db.numero_os)
        nome_arquivo = f"{numero_os_safe}.xlsm"

        caminho_saida = os.path.join(pasta_saida, nome_arquivo)

        # 🔹 Gerar arquivo
        gerar_xlsm(
            modelo="modelos/MODELO_OS.xlsx",
            destino=caminho_saida,
            contexto=contexto,
            mapeamento=MAPEAMENTO_CELULAS
        )

        arquivos_gerados.append(nome_arquivo)

    return {
        "message": "Arquivos gerados com sucesso",
        "total": len(arquivos_gerados),
        "arquivos": arquivos_gerados
    }


@router.get("", response_model=List[OrdemServicoResponse])
def listar_os(
    id_ativo: int | None = None,
    db: Session = Depends(get_db)
):
    query = db.query(OS_models.OrdemServico)

    if id_ativo:
        query = query.filter(
            OS_models.OrdemServico.id_ativo == id_ativo
        )

    return query.order_by(OS_models.OrdemServico.id_os.desc()).all()

@router.get("/ativo/{id_ativo}", response_model=List[OrdemServicoResponse])
def listar_os(
    id_ativo: int,
    db: Session = Depends(get_db)
):
    return (
        db.query(OS_models.OrdemServico)
        .filter(OS_models.OrdemServico.id_ativo == id_ativo)
        .all()
    )


@router.put("/{id_os}")
def editar_ordem_servico(
    id_os: int,
    dados: dict,
    db: Session = Depends(get_db)
):
    campos_somente_leitura = {
        "id_tipo_ativo",
        "tipo_ativo",
        "codigo_ativo",
        "fase",
        "ativo",
        "subestacao",
        "criado_em",
    }

    os_db = (
        db.query(OS_models.OrdemServico)
        .filter(OS_models.OrdemServico.id_os == id_os)
        .first()
    )

    if not os_db:
        raise HTTPException(
            status_code=404,
            detail="Ordem de Serviço não encontrada"
        )


    # Atualiza campos primeiro
    for campo, valor in dados.items():
        if campo in campos_somente_leitura:
            continue
        if hasattr(os_db, campo):
            setattr(os_db, campo, valor)

    # Atualiza instalação baseado na subestação
    if dados.get("id_subestacao"):
        sub = db.query(Subestacao).filter(
            Subestacao.id_subestacao == dados["id_subestacao"]  # ✅ CORRETO
        ).first()

        if not sub:
            raise HTTPException(400, "Subestação inválida")

        os_db.instalacao = sub.nome  # ✅ objeto SQLAlchemy

        


    if os_db.id_ativo:
        ativo = db.query(Ativo).filter(Ativo.id_ativo == os_db.id_ativo).first()
        if ativo:
            os_db.especie = especie_documento_por_ativo(ativo) or os_db.especie

    os_db.prioridade = normalizar_prioridade_operacao(os_db.prioridade)

    status_atualizado = str(os_db.status or "").strip().upper()
    if os_db.numero_ss and status_atualizado in {"ENCERRADA", "CONCLUIDA"}:
        ss_vinculada = db.query(SolicitacaoServico).filter(
            SolicitacaoServico.numero_ss == os_db.numero_ss
        ).first()
        if ss_vinculada:
            ss_vinculada.status = "ENCERRADA"
            ss_vinculada.numero_os = os_db.numero_os

    try:
        db.commit()
        db.refresh(os_db)
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Número da OS já existe"
        )

    return os_db

@router.get("/{id_os}", response_model=OrdemServicoResponse)
def buscar_os_por_id(
    id_os: int,
    db: Session = Depends(get_db)
):
    os = (
        db.query(OS_models.OrdemServico)
        .filter(OS_models.OrdemServico.id_os == id_os)
        .first()
    )

    if not os:
        raise HTTPException(
            status_code=404,
            detail="Os não encontrado"
        )

    return os

@router.delete("/{id_os}")
def deletar_os(
    id_os: int,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin")),
):

    os = db.query(OrdemServico).filter(
        OrdemServico.id_os == id_os
    ).first()

    if not os:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    db.delete(os)
    db.commit()

    return {"message": "OS excluída com sucesso"}


def normalizar_texto_filtro(valor: str | None):
    return str(valor or "").strip().upper()


def chave_ordenacao_codigo(codigo: str | None):
    codigo_normalizado = normalizar_texto_filtro(codigo)
    match = re.search(r"(\d+)", codigo_normalizado)
    if match:
        return (int(match.group(1)), codigo_normalizado)
    return (999999, codigo_normalizado)


def indice_fase(fase: str | None):
    ordem = {
        "AZ": 0,
        "BR": 1,
        "VM": 2,
        "A": 0,
        "B": 1,
        "C": 2,
        "TRIFASICO": 0,
        "NA": 0,
        "N/A": 0,
    }
    return ordem.get(normalizar_texto_filtro(fase), 0)


def derivar_responsaveis_os(
    db: Session,
    id_tipo_ativo: int,
    id_subestacao: int | None,
):
    query = (
        db.query(OS_models.OrdemServico)
        .join(Ativo, Ativo.id_ativo == OS_models.OrdemServico.id_ativo)
        .filter(
            Ativo.id_tipo_ativo == id_tipo_ativo,
            (
                (OS_models.OrdemServico.responsavel_manutencao.isnot(None))
                | (OS_models.OrdemServico.responsavel_operacao.isnot(None))
                | (OS_models.OrdemServico.responsavel.isnot(None))
            ),
        )
    )

    if id_subestacao:
        query = query.filter(OS_models.OrdemServico.id_subestacao == id_subestacao)

    referencia = query.order_by(
        OS_models.OrdemServico.data_fim_execucao.desc(),
        OS_models.OrdemServico.criado_em.desc(),
        OS_models.OrdemServico.id_os.desc(),
    ).first()

    if not referencia:
        return None, None

    return (
        referencia.responsavel_manutencao or referencia.responsavel,
        referencia.responsavel_operacao,
    )


@router.post("/baixa-lote-tipo-ativo", response_model=BaixaOSLoteResponse)
def baixar_os_lote_por_tipo_ativo(
    payload: BaixaOSLoteTipoAtivo,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor")),
):
    if payload.incremento_minutos_por_fase < 0:
        raise HTTPException(
            status_code=400,
            detail="incremento_minutos_por_fase nao pode ser negativo",
        )

    if payload.data_fim_execucao and payload.data_fim_execucao < payload.data_inicio_execucao:
        raise HTTPException(
            status_code=400,
            detail="data_fim_execucao nao pode ser anterior a data_inicio_execucao",
        )

    query = (
        db.query(OS_models.OrdemServico)
        .join(Ativo, Ativo.id_ativo == OS_models.OrdemServico.id_ativo)
        .filter(Ativo.id_tipo_ativo == payload.id_tipo_ativo)
    )

    if payload.id_subestacao:
        query = query.filter(OS_models.OrdemServico.id_subestacao == payload.id_subestacao)

    status_origem = payload.status_origem or ["ABERTA", "PROGRAMADA", "EM_EXECUCAO"]
    if status_origem:
        query = query.filter(OS_models.OrdemServico.status.in_(status_origem))

    if payload.vaos:
        vaos_normalizados = {normalizar_texto_filtro(vao) for vao in payload.vaos}
    else:
        vaos_normalizados = set()

    ordens = query.all()

    if payload.vaos:
        ordens = [
            ordem
            for ordem in ordens
            if normalizar_texto_filtro(ordem.ativo.vao if ordem.ativo else ordem.localizacao)
            in vaos_normalizados
        ]

    if not ordens:
        raise HTTPException(
            status_code=404,
            detail="Nenhuma OS encontrada para baixa com os filtros informados",
        )

    ordens.sort(
        key=lambda ordem: (
            normalizar_texto_filtro(ordem.ativo.vao if ordem.ativo else ordem.localizacao),
            chave_ordenacao_codigo(ordem.ativo.codigo_ativo if ordem.ativo else ordem.codigo_ativo),
            indice_fase(ordem.ativo.fase if ordem.ativo else ordem.complemento),
            ordem.id_os,
        )
    )

    responsavel_manutencao = payload.responsavel_manutencao
    responsavel_operacao = payload.responsavel_operacao

    if payload.derivar_responsaveis and (
        not responsavel_manutencao or not responsavel_operacao
    ):
        manutencao_derivada, operacao_derivada = derivar_responsaveis_os(
            db,
            payload.id_tipo_ativo,
            payload.id_subestacao,
        )
        responsavel_manutencao = responsavel_manutencao or manutencao_derivada
        responsavel_operacao = responsavel_operacao or operacao_derivada

    duracao = None
    if payload.data_fim_execucao:
        duracao = payload.data_fim_execucao - payload.data_inicio_execucao

    por_vao: dict[str, int] = {}
    ordens_atualizadas = []
    indice_por_vao_fase: dict[str, dict[str, int]] = {}

    for ordem in ordens:
        ativo = ordem.ativo
        vao = normalizar_texto_filtro(ativo.vao if ativo else ordem.localizacao) or "SEM_VAO"
        fase = normalizar_texto_filtro(ativo.fase if ativo else ordem.complemento)

        fases_do_vao = indice_por_vao_fase.setdefault(vao, {})
        if fase not in fases_do_vao:
            fases_do_vao[fase] = len(fases_do_vao)

        offset = timedelta(
            minutes=payload.incremento_minutos_por_fase * fases_do_vao[fase]
        )
        inicio_execucao = payload.data_inicio_execucao + offset
        fim_execucao = inicio_execucao + duracao if duracao else inicio_execucao

        ordem.status = payload.status_destino
        ordem.data_inicio_execucao = inicio_execucao
        ordem.data_fim_execucao = fim_execucao

        if responsavel_manutencao:
            ordem.responsavel_manutencao = responsavel_manutencao
        if responsavel_operacao:
            ordem.responsavel_operacao = responsavel_operacao

        if payload.observacao_baixa:
            observacoes = ordem.observacoes or ""
            separador = "\n" if observacoes else ""
            ordem.observacoes = f"{observacoes}{separador}{payload.observacao_baixa}"

        if ordem.numero_ss and str(payload.status_destino or "").strip().upper() in {
            "ENCERRADA",
            "CONCLUIDA",
        }:
            ss_vinculada = db.query(SolicitacaoServico).filter(
                SolicitacaoServico.numero_ss == ordem.numero_ss
            ).first()
            if ss_vinculada:
                ss_vinculada.status = "ENCERRADA"
                ss_vinculada.numero_os = ordem.numero_os

        por_vao[vao] = por_vao.get(vao, 0) + 1
        ordens_atualizadas.append(ordem)

    db.commit()

    for ordem in ordens_atualizadas:
        db.refresh(ordem)

    return {
        "mensagem": "Baixa em lote realizada com sucesso",
        "total": len(ordens_atualizadas),
        "por_vao": por_vao,
        "ordens": [
            {
                "id_os": ordem.id_os,
                "numero_os": ordem.numero_os,
                "codigo_ativo": ordem.ativo.codigo_ativo if ordem.ativo else None,
                "vao": ordem.ativo.vao if ordem.ativo else ordem.localizacao,
                "fase": ordem.ativo.fase if ordem.ativo else ordem.complemento,
                "status": ordem.status,
                "data_inicio_execucao": ordem.data_inicio_execucao,
                "data_fim_execucao": ordem.data_fim_execucao,
                "responsavel_manutencao": ordem.responsavel_manutencao,
                "responsavel_operacao": ordem.responsavel_operacao,
            }
            for ordem in ordens_atualizadas
        ],
    }


@router.post("/lote-por-tipo-ativo", response_model=List[OrdemServicoResponse])
def criar_os_lote_por_tipo_ativo(
    payload: OrdemServicoCreateLote,
    db: Session = Depends(get_db),
):
    # ====================== 1. VALIDAÇÃO E BUSCA DA SUBESTAÇÃO ======================
    if not payload.id_subestacao:
        raise HTTPException(status_code=400, detail="id_subestacao é obrigatório")

    subestacao = db.query(Subestacao).filter(
        Subestacao.id_subestacao == payload.id_subestacao
    ).first()

    if not subestacao:
        raise HTTPException(status_code=400, detail="Subestação inválida")

    instalacao_nome = subestacao.nome   # ← Aqui pegamos o nome real da subestação

    codigo_ativo_filtro = str(payload.codigo_ativo or "").strip()
    if not payload.id_tipo_ativo and not codigo_ativo_filtro:
        raise HTTPException(
            status_code=400,
            detail="Informe id_tipo_ativo ou codigo_ativo para gerar OS em lote.",
        )

    # ====================== 2. BUSCAR ATIVOS ======================
    query_ativos = db.query(Ativo).filter(
        Ativo.id_subestacao == payload.id_subestacao,
    )

    if payload.id_tipo_ativo:
        query_ativos = query_ativos.filter(Ativo.id_tipo_ativo == payload.id_tipo_ativo)

    if codigo_ativo_filtro:
        query_ativos = query_ativos.filter(
            text("UPPER(TRIM(codigo_ativo)) = UPPER(TRIM(:codigo_ativo))")
        ).params(codigo_ativo=codigo_ativo_filtro)

    ativos = query_ativos.all()

    if codigo_ativo_filtro:
        fases_desejadas = {"AZ", "BR", "VM"}
        if payload.incluir_reserva:
            fases_desejadas.update({"RES", "RESERVA"})

        ativos = [
            ativo
            for ativo in ativos
            if str(ativo.fase or "").strip().upper() in fases_desejadas
        ]

    if not ativos:
        raise HTTPException(
            status_code=404,
            detail="Nenhum ativo encontrado para esse tipo na subestação."
        )


 
    # 🔹 Definir sigla (baseado na lista)
    try:
        sigla = SUBESTACOES_SIGLAS[payload.id_subestacao - 1]
    except IndexError:
        raise HTTPException(400, "Subestação sem sigla configurada")
    


    # ====================== 3. ORDENAÇÃO CORRETA ======================
    ordem_fase = {"AZ": 1, "BR": 2, "VM": 3}

    def normalizar_fase(fase: str | None) -> str:
        return str(fase).strip().upper() if fase else ""

    def ordenar_codigo_ativo(codigo: str | None) -> tuple:
        if not codigo:
            return (999, "")
        codigo_str = str(codigo).strip().upper()
        match = re.match(r"(\d+)(.*)", codigo_str)
        if match:
            num = int(match.group(1))
            resto = match.group(2)
            return (num, resto)
        return (999, codigo_str)

    ativos.sort(
        key=lambda a: (
            ordenar_codigo_ativo(a.codigo_ativo),
            ordem_fase.get(normalizar_fase(a.fase), 99)
        )
    )

    # ====================== 4. GERAÇÃO DO NÚMERO OS ======================
 
   # 🔹 Validar Ativo
    ativo = None
    codigo_ativo = None

   

    payload.numero_os,  payload.numero_apr = gerar_numero_os(db, sigla, codigo_ativo)
    
  

    padrao = r"^(.*-)(\d+)-(\d{4})$"
    match = re.match(padrao, payload.numero_os.strip())
    match2 = re.match(padrao, payload.numero_apr.strip())
    if not match:
        raise HTTPException(
            status_code=400,
            detail="Formato inválido. Use: OS-RTV-BJD-0168-2026"
        )

    prefixo = match.group(1)
    prefixo2 = match2.group(1)

    numero_base = int(match.group(2))
    ano = match.group(3)
    padding = len(match.group(2))

    ordens_criadas = []
    

    # ====================== 5. CRIAR AS OS ======================
    for index, ativo in enumerate(ativos):
        numero_atual = numero_base + index
        numero_formatado = str(numero_atual).zfill(padding)

        numero_os_final = f"{prefixo}{numero_formatado}-{ano}-{ativo.codigo_ativo}"
        numero_apr_final = f"{prefixo2}{numero_formatado}-{ano}"

        fase = normalizar_fase(ativo.fase)
        complemento = f"Fase: {fase}"
        local =  f"{ativo.vao}"

        nova_os = OS_models.OrdemServico(
            numero_os=numero_os_final,
            id_subestacao=payload.id_subestacao,
            id_ativo=ativo.id_ativo,
            especie=especie_documento_por_ativo(ativo) or payload.especie or getattr(payload, "tipo_ativo", None),
            numero_si=payload.numero_si,
            numero_apr=numero_apr_final,

            # ✅ CORRIGIDO: Usa o nome da subestação real
            instalacao=instalacao_nome,
            localizacao=local,
            complemento=complemento,

            origens=payload.origens,
            defeito=payload.defeito,
            esquema_servicos=payload.esquema_servicos,
            causa_primaria=payload.causa_primaria,
            causa_secundaria=payload.causa_secundaria,

            prioridade=normalizar_prioridade_operacao(payload.prioridade),
            responsavel=payload.responsavel,
            responsavel_manutencao=payload.responsavel_manutencao,
            responsavel_operacao=payload.responsavel_operacao,
            substituto=payload.substituto,

            data_abertura_ss=payload.data_abertura_ss,
            data_inicio_programado=payload.data_inicio_programado,
            data_fim_programado=payload.data_fim_programado,

            descricao_servicos=payload.descricao_servicos,
            observacoes=payload.observacoes,
            centro_custos=payload.centro_custos or "RIALMA TRANSMISSORA V",
            status=payload.status or "ABERTA",
            emissor=payload.emissor,
        )

        try:
            db.add(nova_os)
            db.flush()

            # Gerar XLSM
            ativo_obj = db.query(Ativo).get(ativo.id_ativo)
            contexto = montar_contexto_os(nova_os, ativo_obj)

            pasta_saida = "saida"
            os.makedirs(pasta_saida, exist_ok=True)

            caminho_saida = os.path.join(pasta_saida, f"{nome_arquivo_seguro(nova_os.numero_os)}.xlsm")

            gerar_xlsm(
                modelo="modelos/MODELO_OS.xlsx",
                destino=caminho_saida,
                contexto=contexto,
                mapeamento=MAPEAMENTO_CELULAS
            )

            ordens_criadas.append(nova_os)

        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Erro ao criar OS para {ativo.codigo_ativo} - Fase {fase}: {str(e)}"
            )

    db.commit()
    return ordens_criadas


from models.plano_manutencao_models import (
    PlanoItem,
    PlanoExecucao,
    PeriodicidadeEnum,
    PlanoManutencao

)

from models.familias_models import (
    TipoAtivo

)


from datetime import datetime, time, timedelta
from dateutil.relativedelta import relativedelta

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session

from database import get_db




def delta_periodicidade(periodicidade: PeriodicidadeEnum, intervalo: int | None = 1):
    multiplicador = max(intervalo or 1, 1)

    if periodicidade == PeriodicidadeEnum.SEMANAL:
        return relativedelta(weeks=multiplicador)
    if periodicidade == PeriodicidadeEnum.MENSAL:
        return relativedelta(months=multiplicador)
    if periodicidade == PeriodicidadeEnum.BIMESTRAL:
        return relativedelta(months=2 * multiplicador)
    if periodicidade == PeriodicidadeEnum.TRIMESTRAL:
        return relativedelta(months=3 * multiplicador)
    if periodicidade == PeriodicidadeEnum.SEMESTRAL:
        return relativedelta(months=6 * multiplicador)
    if periodicidade == PeriodicidadeEnum.TRES_ANOS:
        return relativedelta(years=3 * multiplicador)
    if periodicidade == PeriodicidadeEnum.CINCO_ANOS:
        return relativedelta(years=5 * multiplicador)
    if periodicidade == PeriodicidadeEnum.SEIS_ANOS:
        return relativedelta(years=6 * multiplicador)

    return relativedelta(weeks=multiplicador)


def data_inicial_execucao(item: PlanoItem, hoje: datetime):
    if item.data_inicio:
        return datetime.combine(item.data_inicio, time.min)

    return hoje


def deve_gerar_os(item: PlanoItem, proxima_execucao: datetime, hoje: datetime):
    antecedencia = max(item.antecedencia or 0, 0)
    data_liberacao = proxima_execucao - timedelta(days=antecedencia)

    return data_liberacao <= hoje


def proxima_data_execucao(item: PlanoItem, data_atual: datetime, hoje: datetime):
    proxima = data_atual + delta_periodicidade(item.periodicidade, item.intervalo)

    while deve_gerar_os(item, proxima, hoje):
        proxima += delta_periodicidade(item.periodicidade, item.intervalo)

    return proxima


def valor_periodicidade(periodicidade: PeriodicidadeEnum):
    return getattr(periodicidade, "value", periodicidade)


def data_programada_os(execucoes_pendentes: list[tuple[PlanoItem, PlanoExecucao]], hoje: datetime):
    primeira_data_vencida = min(
        execucao.proxima_execucao or hoje
        for _, execucao in execucoes_pendentes
    )

    if primeira_data_vencida < hoje:
        return hoje

    return primeira_data_vencida


def gerar_os_por_planos_manutencao(db: Session):
    hoje = datetime.now()
    os_criadas = []

    tipos_ativo = db.query(TipoAtivo).all()

    for tipo in tipos_ativo:
        planos = (
            db.query(PlanoManutencao)
            .filter(PlanoManutencao.id_tipo_ativo == tipo.id_tipo_ativo)
            .all()
        )

        for plano in planos:
            itens_plano = (
                db.query(PlanoItem)
                .filter(PlanoItem.id_plano_manutencao == plano.id_plano_manutencao)
                .all()
            )

            if not itens_plano:
                continue

            ativos = (
                db.query(Ativo)
                .filter(Ativo.id_tipo_ativo == tipo.id_tipo_ativo)
                .all()
            )

            for ativo in ativos:
                execucoes_pendentes = []

                for item in itens_plano:
                    execucao = (
                        db.query(PlanoExecucao)
                        .filter(
                            PlanoExecucao.id_plano_item == item.id_plano_item,
                            PlanoExecucao.id_ativo == ativo.id_ativo,
                        )
                        .first()
                    )

                    if not execucao:
                        execucao = PlanoExecucao(
                            id_plano_item=item.id_plano_item,
                            id_ativo=ativo.id_ativo,
                            ultima_execucao=None,
                            proxima_execucao=data_inicial_execucao(item, hoje),
                        )
                        db.add(execucao)
                        db.flush()

                    if execucao.proxima_execucao and not deve_gerar_os(
                        item,
                        execucao.proxima_execucao,
                        hoje,
                    ):
                        continue

                    execucoes_pendentes.append((item, execucao))

                if not execucoes_pendentes:
                    continue

                data_programada = data_programada_os(execucoes_pendentes, hoje)

                os_existente = (
                    db.query(OrdemServico)
                    .filter(
                        OrdemServico.id_ativo == ativo.id_ativo,
                        OrdemServico.descricao_servicos == plano.descricao_geral,
                        OrdemServico.status.in_(["ABERTA", "PROGRAMADA", "EM_EXECUCAO"]),
                    )
                    .first()
                )

                if os_existente:
                    for item, execucao in execucoes_pendentes:
                        execucao.ultima_execucao = hoje
                        execucao.proxima_execucao = proxima_data_execucao(
                            item,
                            execucao.proxima_execucao or hoje,
                            hoje,
                        )
                    continue

                subestacao = (
                    db.query(Subestacao)
                    .filter(Subestacao.id_subestacao == ativo.id_subestacao)
                    .first()
                )

                if not subestacao:
                    continue

                try:
                    sigla = SUBESTACOES_SIGLAS[ativo.id_subestacao - 1]
                except IndexError:
                    continue

                numero_os, numero_apr = gerar_numero_os(
                    db,
                    sigla,
                    ativo.codigo_ativo,
                )

                ultima_os_com_equipe = (
                    db.query(OrdemServico)
                    .filter(
                        OrdemServico.id_subestacao == ativo.id_subestacao,
                        OrdemServico.responsavel.isnot(None),
                        OrdemServico.responsavel != "",
                        OrdemServico.substituto.isnot(None),
                        OrdemServico.substituto != "",
                    )
                    .order_by(
                        OrdemServico.criado_em.desc(),
                        OrdemServico.id_os.desc(),
                    )
                    .first()
                )

                nova_os = OrdemServico(
                    numero_os=numero_os,
                    numero_apr=numero_apr,
                    id_subestacao=ativo.id_subestacao,
                    id_ativo=ativo.id_ativo,
                    especie=especie_documento_por_ativo(ativo),
                    instalacao=subestacao.nome,
                    localizacao=ativo.vao,
                    complemento=ativo.fase,
                    descricao_servicos=plano.descricao_geral,
                    data_inicio_programado=data_programada,
                    data_fim_programado=data_programada,
                    responsavel=ultima_os_com_equipe.responsavel if ultima_os_com_equipe else None,
                    substituto=ultima_os_com_equipe.substituto if ultima_os_com_equipe else None,
                    prioridade="NIVEL_3",
                    status="ABERTA",
                )

                db.add(nova_os)
                db.flush()

                for item, execucao in execucoes_pendentes:
                    execucao.ultima_execucao = hoje
                    execucao.proxima_execucao = proxima_data_execucao(
                        item,
                        execucao.proxima_execucao or hoje,
                        hoje,
                    )

                os_criadas.append({
                    "numero_os": numero_os,
                    "ativo": ativo.codigo_ativo,
                    "itens_plano": [
                        {
                            "nome_item": item.nome_item,
                            "periodicidade": valor_periodicidade(item.periodicidade),
                        }
                        for item, _ in execucoes_pendentes
                    ],
                    "responsavel": nova_os.responsavel,
                    "substituto": nova_os.substituto,
                })

    db.commit()

    return {
        "mensagem": "OS preventivas geradas com sucesso",
        "total_os": len(os_criadas),
        "os_criadas": os_criadas,
    }


@router.post("/gerar-os-planos")
def gerar_os_planos(db: Session = Depends(get_db)):
    return gerar_os_por_planos_manutencao(db)


@router.post("/gerar-os-semanal")
def gerar_os_semanal(db: Session = Depends(get_db)):
    return gerar_os_por_planos_manutencao(db)
