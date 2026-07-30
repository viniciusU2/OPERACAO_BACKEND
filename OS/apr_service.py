
from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from database import Base, engine
from models.APR_models import APR, APRHistorico, FrenteServico, FrenteServicoOS
from models.Ativo import Ativo
from models.OS_models import OrdemServico
from models.instalacao_models import Subestacao


MODELO_APR = "modelos/MODELO_APR.xlsm"


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    return str(valor)


def nome_arquivo_seguro(texto: str | None) -> str:
    if not texto:
        return "sem_nome"
    return re.sub(r"[^A-Za-z0-9_-]", "_", texto)


def garantir_estrutura_apr(db: Session):
    Base.metadata.create_all(bind=engine)

    existe_coluna = db.execute(
        text("SHOW COLUMNS FROM ordem_servico LIKE 'id_frente_servico'")
    ).first()
    if not existe_coluna:
        db.execute(text("ALTER TABLE ordem_servico ADD COLUMN id_frente_servico INT NULL"))
        db.commit()


def _proximo_codigo_frente(db: Session, sigla: str) -> str:
    ano = datetime.now().year
    registros = (
        db.query(FrenteServico.codigo_frente)
        .filter(FrenteServico.codigo_frente.like(f"FS-{sigla}-%-{ano}"))
        .all()
    )
    numeros: list[int] = []
    for (codigo,) in registros:
        match = re.search(rf"FS-{re.escape(sigla)}-(\d+)-{ano}", codigo or "")
        if match:
            numeros.append(int(match.group(1)))
    return f"FS-{sigla}-{str((max(numeros) + 1) if numeros else 1).zfill(4)}-{ano}"


def sigla_por_os(os_db: OrdemServico) -> str:
    numero = os_db.numero_os or os_db.numero_apr or ""
    match = re.search(r"(?:OS|APR)-([A-Z]+)-\d+-\d{4}", numero)
    if match:
        return match.group(1)
    if os_db.instalacao:
        partes = str(os_db.instalacao).split()
        if partes:
            return partes[-1].upper()
    return "APR"


def formatar_periodo(ordens: list[OrdemServico]) -> str:
    datas_inicio = [o.data_inicio_programado for o in ordens if o.data_inicio_programado]
    datas_fim = [o.data_fim_programado for o in ordens if o.data_fim_programado]
    inicio = min(datas_inicio) if datas_inicio else None
    fim = max(datas_fim) if datas_fim else None

    if inicio and fim:
        if inicio.date() == fim.date():
            return f"{inicio.strftime('%d/%m/%y')} das {inicio.strftime('%H:%M')} as {fim.strftime('%H:%M')}"
        return f"{inicio.strftime('%d/%m/%y %H:%M')} a {fim.strftime('%d/%m/%y %H:%M')}"
    return limpar(inicio or fim)


def formatar_lista_os(ordens: list[OrdemServico]) -> str:
    numeros = [o.numero_os for o in ordens if o.numero_os]
    if not numeros:
        return ""

    parsed = []
    for numero in numeros:
        match = re.match(r"^(.*-)(\d+)-(\d{4})$", numero)
        if not match:
            parsed = []
            break
        parsed.append((numero, match.group(1), int(match.group(2)), match.group(3), len(match.group(2))))

    if len(parsed) > 1:
        prefixos = {p[1] for p in parsed}
        anos = {p[3] for p in parsed}
        sequencias = sorted(p[2] for p in parsed)
        if len(prefixos) == 1 and len(anos) == 1 and sequencias == list(range(sequencias[0], sequencias[-1] + 1)):
            prefixo = parsed[0][1]
            ano = parsed[0][3]
            padding = parsed[0][4]
            primeiro = f"{prefixo}{str(sequencias[0]).zfill(padding)}-{ano}"
            ultimo = f"{prefixo}{str(sequencias[-1]).zfill(padding)}-{ano}"
            return f"{primeiro} ate {ultimo}"

    return ", ".join(numeros)


def _ativos_responsaveis(db: Session, ordens: list[OrdemServico]) -> tuple[str, str]:
    responsaveis: list[str] = []
    substitutos: list[str] = []

    for ordem in ordens:
        responsavel = ordem.responsavel or ordem.responsavel_manutencao or ""
        if responsavel and responsavel not in responsaveis:
            responsaveis.append(responsavel)
        if ordem.substituto and ordem.substituto not in substitutos:
            substitutos.append(ordem.substituto)

    return "\n".join(responsaveis), "\n".join(f"Substituto: {nome}" for nome in substitutos)


def registrar_historico_apr(db: Session, apr: APR, acao: str, usuario: str | None = None, observacao: str | None = None):
    db.add(APRHistorico(id_apr=apr.id_apr, acao=acao, usuario=usuario, observacao=observacao))


def criar_frente_para_ordens(
    db: Session,
    ordens: Iterable[OrdemServico],
    origem: str,
    numero_apr: str | None = None,
    usuario: str | None = None,
) -> FrenteServico | None:
    garantir_estrutura_apr(db)
    ordens_lista = [ordem for ordem in ordens if ordem]
    if not ordens_lista:
        return None

    ids_frentes = {getattr(ordem, "id_frente_servico", None) for ordem in ordens_lista if getattr(ordem, "id_frente_servico", None)}
    if ids_frentes:
        frente_existente = db.query(FrenteServico).filter(FrenteServico.id_frente_servico == list(ids_frentes)[0]).first()
        if frente_existente:
            return frente_existente

    referencia = ordens_lista[0]
    sigla = sigla_por_os(referencia)
    periodo_inicio = min([o.data_inicio_programado for o in ordens_lista if o.data_inicio_programado], default=None)
    periodo_fim = max([o.data_fim_programado for o in ordens_lista if o.data_fim_programado], default=None)
    descricao = referencia.descricao_servicos or referencia.defeito or referencia.esquema_servicos

    frente = FrenteServico(
        codigo_frente=_proximo_codigo_frente(db, sigla),
        origem=origem,
        id_subestacao=referencia.id_subestacao,
        sigla_subestacao=sigla,
        descricao_atividade=descricao,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        responsavel=referencia.responsavel or referencia.responsavel_manutencao,
        substituto=referencia.substituto,
        status="ABERTA",
        criado_por=usuario,
    )
    db.add(frente)
    db.flush()

    numero_apr_final = numero_apr or referencia.numero_apr or f"APR-{sigla}-{str(frente.id_frente_servico).zfill(4)}-{datetime.now().year}"
    apr = APR(
        id_frente_servico=frente.id_frente_servico,
        numero_apr=numero_apr_final,
        gerada_por=usuario,
        modelo_versao=os.path.basename(MODELO_APR),
    )
    db.add(apr)
    db.flush()

    for index, ordem in enumerate(ordens_lista, start=1):
        ordem.id_frente_servico = frente.id_frente_servico
        ordem.numero_apr = numero_apr_final
        db.add(FrenteServicoOS(id_frente_servico=frente.id_frente_servico, id_os=ordem.id_os, ordem=index))

    registrar_historico_apr(db, apr, "CRIADA", usuario, f"APR criada para {len(ordens_lista)} OS")
    return frente


def obter_frente_por_os(db: Session, os_db: OrdemServico, criar_se_nao_existir: bool = True) -> FrenteServico | None:
    garantir_estrutura_apr(db)
    if getattr(os_db, "id_frente_servico", None):
        frente = db.query(FrenteServico).filter(FrenteServico.id_frente_servico == os_db.id_frente_servico).first()
        if frente:
            return frente
    if criar_se_nao_existir:
        frente = criar_frente_para_ordens(db, [os_db], "MANUAL_LEGADO", os_db.numero_apr)
        db.commit()
        return frente
    return None


def ordens_da_frente(db: Session, frente: FrenteServico) -> list[OrdemServico]:
    vinculos = (
        db.query(FrenteServicoOS)
        .filter(FrenteServicoOS.id_frente_servico == frente.id_frente_servico)
        .order_by(FrenteServicoOS.ordem.asc(), FrenteServicoOS.id.asc())
        .all()
    )
    ids = [v.id_os for v in vinculos]
    if not ids:
        return []
    ordens = db.query(OrdemServico).filter(OrdemServico.id_os.in_(ids)).all()
    por_id = {ordem.id_os: ordem for ordem in ordens}
    return [por_id[id_os] for id_os in ids if id_os in por_id]


def gerar_apr_xlsm(db: Session, frente: FrenteServico, destino: str) -> str:
    apr = frente.apr or db.query(APR).filter(APR.id_frente_servico == frente.id_frente_servico).first()
    ordens = ordens_da_frente(db, frente)
    if not apr or not ordens:
        raise ValueError("Frente de servi?o sem APR ou OS vinculada")

    shutil.copy(MODELO_APR, destino)
    wb = load_workbook(destino, keep_vba=True)
    ws = wb.active

    ativos_responsaveis, substitutos = _ativos_responsaveis(db, ordens)
    ws["D4"] = frente.descricao_atividade or ordens[0].descricao_servicos or ordens[0].defeito or ""
    ws["D7"] = f"SE {frente.sigla_subestacao}" if frente.sigla_subestacao else limpar(ordens[0].instalacao)
    ws["G7"] = frente.responsavel or ordens[0].responsavel or ordens[0].responsavel_manutencao or ""
    ws["L6"] = f"5. APR / N? {apr.numero_apr}\nOS {formatar_lista_os(ordens)}"
    ws["O7"] = formatar_periodo(ordens)
    ws["C14"] = ativos_responsaveis
    ws["C15"] = substitutos

    ws["C14"].alignment = ws["C14"].alignment.copy(wrap_text=True, vertical="top")
    ws["C15"].alignment = ws["C15"].alignment.copy(wrap_text=True, vertical="top")
    ws.row_dimensions[14].height = max(12, min(80, 12 * max(1, len(ativos_responsaveis.splitlines()))))
    ws.row_dimensions[15].height = max(12, min(60, 12 * max(1, len(substitutos.splitlines()))))

    wb.save(destino)
    apr.caminho_arquivo = destino
    apr.status = "GERADA"
    apr.atualizado_em = datetime.utcnow()
    registrar_historico_apr(db, apr, "GERADA", apr.gerada_por, "Arquivo APR gerado")
    return destino
