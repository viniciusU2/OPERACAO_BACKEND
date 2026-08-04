from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import text
from sqlalchemy.orm import Session, selectinload
from fastapi import Query
from sqlalchemy import or_
from auth.dependencies import get_current_user, require_roles
from database import get_db
from models.SI_models import SILiberacao, solicitacao_intervencao
from models.instalacao_models import Subestacao
from models.Ativo import Ativo, GrupoAtivo
from ATIVO.grupos_ativos import garantir_estrutura_grupo_ativo, sincronizar_grupos_ativos, validar_selecao_ativo
from SI.schemas import (
    SICreate,
    SILiberacaoCancelarUpdate,
    SILiberacaoManutencaoCreate,
    SILiberacaoOperacaoUpdate,
    SILiberacaoResponse,
    SIResponse,
    SIPaginadaResponse,
    SIUpdate,
)
from utils.documentos_operacao import (
    especie_documento_por_ativo,
    normalizar_prioridade_operacao,
)

import os
import shutil
import re
from datetime import date, datetime

from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries

router = APIRouter(prefix="/si", tags=["Servi?o Interven??o"])

SUBESTACOES_SIGLAS = ["BJD", "GOR", "JAB"]


def garantir_colunas_si(db: Session):
    colunas = {
        "numero_os": "VARCHAR(30) NULL",
        "natureza": "VARCHAR(255) NULL",
        "caracteristica_intervencao": "VARCHAR(100) NULL",
        "prioridade": "VARCHAR(20) NULL DEFAULT 'NIVEL_3'",
        "risco_desligamento": "TEXT NULL",
        "condicoes_climaticas": "TEXT NULL",
        "execucao_periodo_noturno": "TEXT NULL",
        "postergacao_traz_risco": "VARCHAR(30) NULL",
        "acarreta_risco_perdas_multiplas": "VARCHAR(30) NULL DEFAULT 'NAO'",
        "editado_por": "TEXT NULL",
        "quais_risco_desligamento": "TEXT NULL",
        "quais_condicoes_climaticas": "TEXT NULL",
        "quais_execucao_periodo_noturno": "TEXT NULL",
    }

    for coluna, definicao in colunas.items():
        existe = db.execute(
            text("SHOW COLUMNS FROM solicitacao_intervencao LIKE :coluna"),
            {"coluna": coluna},
        ).first()
        if not existe:
            db.execute(text(f"ALTER TABLE solicitacao_intervencao ADD COLUMN {coluna} {definicao}"))

    db.commit()


def garantir_tabela_liberacoes_si(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS si_liberacao (
            id_liberacao INT AUTO_INCREMENT PRIMARY KEY,
            id_si INT NOT NULL,
            data_utilizacao DATE NOT NULL,
            data_hora_liberacao DATETIME NOT NULL,
            usuario_solicitou_id INT NULL,
            usuario_solicitou VARCHAR(200) NOT NULL,
            operador_liberou VARCHAR(200) NULL,
            data_hora_devolucao DATETIME NULL,
            usuario_devolveu_id INT NULL,
            usuario_devolveu VARCHAR(200) NULL,
            operador_recebeu_devolucao VARCHAR(200) NULL,
            observacoes TEXT NULL,
            status VARCHAR(30) NOT NULL DEFAULT 'EM_EXECUCAO',
            criado_em DATETIME NULL,
            atualizado_em DATETIME NULL,
            INDEX idx_si_liberacao_id_si (id_si),
            INDEX idx_si_liberacao_status (status),
            CONSTRAINT fk_si_liberacao_si
                FOREIGN KEY (id_si) REFERENCES solicitacao_intervencao(id_si)
                ON DELETE CASCADE,
            CONSTRAINT fk_si_liberacao_usuario_solicitou
                FOREIGN KEY (usuario_solicitou_id) REFERENCES usuarios(id),
            CONSTRAINT fk_si_liberacao_usuario_devolveu
                FOREIGN KEY (usuario_devolveu_id) REFERENCES usuarios(id)
        )
    """))
    db.commit()


def nome_arquivo_seguro(texto: str):
    if not texto:
        return "sem_nome"

    return re.sub(r'[^A-Za-z0-9_.-]', '_', texto)


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def formatar_risco_postergacao(valor):
    valores = {
        "NAO": "NÃO",
        "SIM_EQUIPAMENTO": "SIM - Equipamento",
        "SIM_PESSOA": "SIM - Pessoa",
    }
    valor_normalizado = str(valor or "NAO").strip().upper()
    return valores.get(valor_normalizado, limpar(valor))


def set_valor_seguro(ws, celula, valor):
    for merged_range in ws.merged_cells.ranges:
        if celula in merged_range:
            min_col, min_row, _, _ = range_boundaries(str(merged_range))
            ws.cell(row=min_row, column=min_col).value = valor
            return
    ws[celula] = valor


def sigla_por_subestacao(id_subestacao: int | None):
    if not id_subestacao:
        return "GERAL"

    try:
        return SUBESTACOES_SIGLAS[id_subestacao - 1]
    except IndexError:
        return "GERAL"


def gerar_numero_si(db: Session, sigla: str):
    ano_atual = datetime.now().year
    registros = (
        db.query(solicitacao_intervencao.numero_si)
        .filter(solicitacao_intervencao.numero_si.like(f"%{sigla}%{ano_atual}%"))
        .all()
    )

    numeros = []
    for (numero_si,) in registros:
        numero_si = numero_si or ""
        if sigla not in numero_si:
            continue
        match = re.search(rf"(\d+)-{ano_atual}", numero_si)
        if match:
            numeros.append(int(match.group(1)))

    proximo = max(numeros) + 1 if numeros else 1
    return f"SI-{sigla}-{str(proximo).zfill(4)}-{ano_atual}"


def obter_si_ou_404(db: Session, id_si: int):
    si = db.query(solicitacao_intervencao).filter(
        solicitacao_intervencao.id_si == id_si
    ).first()
    if not si:
        raise HTTPException(404, "SI n?o encontrada")
    return si


def nome_usuario(usuario):
    return getattr(usuario, "nome", None) or getattr(usuario, "email", None) or "Usu?rio"


MAPEAMENTO_CELULAS = {
    "NUM_SI": "H1",
    "NUM_SGI": "H3",
    "NUM_APR": "A5",
    "NUM_OS": "C5",
    "ESPECIE": "G5",
    "INSTALACAO": "A7",
    "LOCALIZACAO": "D7",
    "CODIGO_ATIVO": "G7",
    "NATUREZA": "A9",
    "CARACTERISTICA_INTERVENCAO": "D9",
    "TIPO": "G9",
    "DOCUMENTOS_REFERENCIA": "A11",
    "DT_INICIO_PRERIODO_TOTAL": "A13",
    "DT_FIM_PRERIODO_TOTAL": "F13",
    "DT_INICIO_PRERIODO_MANUTENCAO": "A15",
    "DT_FIM_PRERIODO_MANUTENCAO": "F15",
    "JUSTIFICATIVA": "A17",
    "RESPONSAVEL": "A19",
    "SUBSTITUTO": "F19",
    "APROVEITAMENTO": "A21",
    "INCLUSAO_SERVICO": "C21",
    "ACARRETA_RISCO_PERDAS_MULTIPLAS": "E21",
    "POSTERGACAO_TRAZ_RISCO": "G21",
    "ORGAOS": "I21",
    "TIPO_PROGRAMACAO": "A24",
    "DIAS_EXCECAO": "C24",
    "TEMPO_RETORNO": "F24",
    "DESC_SERVICOS": "A27",
    "OBSERVACOES": "A29",
    "CABO_ATERRAMENTO": "A31",
    "RISCO_DESLIGAMENTO": "A33",
    "QUAIS_RISCO_DESLIGAMENTO": "C33",
    "CONDICOES_CLIMATICAS": "A36",
    "QUAIS_CONDICOES_CLIMATICAS": "C36",
    "EXECUCAO_PERIODO_NOTURNO": "A39",
    "QUAIS_EXECUCAO_PERIODO_NOTURNO": "C39",
    "RESPONSAVEL_MANUTENCAO_ONS": "B42",
    "RESPONSAVEL_MANUTENCAO_COT": "E42",
    "RESPONSAVEL_MANUTENCAO_SE": "H42",
    "RESPONSAVEL_DATA_MANUTENCAO_ONS": "B43",
    "RESPONSAVEL_DATA_MANUTENCAO_COT": "E43",
    "RESPONSAVEL_DATA_MANUTENCAO_SE": "H43",
    "ASSINATURA_MANUTENCAO": "D44",
    "STATUS_MANUTENCAO": "H44",
    "DATA_MANUTENCAO": "D45",
    "RESPONSAVEL_OPERACAO_SE": "B52",
    "RESPONSAVEL_OPERACAO_COT": "E52",
    "RESPONSAVEL_OPERACAO_ONS": "H52",
    "ASSINATURA_OPERACAO": "D50",
    "STATUS_OPERACAO": "H50",
    "DATA_OPERACAO": "D51",
}


def adicionar_aba_liberacoes(wb, liberacoes, contexto=None):
    if "Liberacoes" in wb.sheetnames:
        del wb["Liberacoes"]

    contexto = contexto or {}
    ws = wb.create_sheet("Liberacoes")
    cinza = PatternFill("solid", fgColor="BFBFBF")
    branco = PatternFill("solid", fgColor="FFFFFF")
    borda_fina = Side(style="thin", color="000000")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    centro = Alignment(horizontal="center", vertical="center")
    esquerda = Alignment(horizontal="left", vertical="center")

    larguras = [13.0, 10.4, 13.0, 13.0, 13.0, 9.3, 12.0, 9.6, 13.9]
    for index, largura in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = largura

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.511811024
    ws.page_margins.right = 0.511811024
    ws.page_margins.top = 0.787401575
    ws.page_margins.bottom = 0.787401575

    def preparar_linha(row, height=20):
        ws.row_dimensions[row].height = height
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = borda
            cell.fill = branco
            cell.alignment = esquerda
            cell.font = Font(name="Arial", size=10)

    def titulo(row, texto):
        preparar_linha(row, 20)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1)
        cell.value = texto
        cell.fill = cinza
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = centro

    def bloco_label(row, col, texto, col_fim=None):
        col_fim = col_fim or col
        if col_fim > col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_fim)
        cell = ws.cell(row=row, column=col)
        cell.value = texto
        cell.fill = cinza
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = esquerda

    def bloco_valor(row, col, valor, col_fim=None, center=False):
        col_fim = col_fim or col
        if col_fim > col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_fim)
        cell = ws.cell(row=row, column=col)
        cell.value = valor
        cell.alignment = centro if center else esquerda

    def render_liberacao(liberacao, row):
        status = {"EM_EXECUCAO": "EM EXECUCAO", "ENCERRADA": "ENCERRADA", "CANCELADA": "CANCELADA", "ABERTA": "ABERTA"}.get(liberacao.status, limpar(liberacao.status)).upper()

        titulo(row, "Libera\u00e7\u00e3o para manuten\u00e7\u00e3o")

        for offset in range(1, 5):
            preparar_linha(row + offset)

        bloco_label(row + 1, 1, "ONS:")
        bloco_valor(row + 1, 2, contexto.get("RESPONSAVEL_MANUTENCAO_ONS", ""), 3)
        bloco_label(row + 1, 4, "COT:")
        bloco_valor(row + 1, 5, contexto.get("RESPONSAVEL_MANUTENCAO_COT", ""), 6)
        bloco_label(row + 1, 7, "SE:")
        bloco_valor(row + 1, 8, contexto.get("RESPONSAVEL_MANUTENCAO_SE", ""), 9)

        bloco_label(row + 2, 1, "Data/Hora")
        bloco_valor(row + 2, 2, contexto.get("RESPONSAVEL_DATA_MANUTENCAO_ONS", ""), 3)
        bloco_label(row + 2, 4, "Data/Hora")
        bloco_valor(row + 2, 5, contexto.get("RESPONSAVEL_DATA_MANUTENCAO_COT", ""), 6)
        bloco_label(row + 2, 7, "Data/Hora:")
        bloco_valor(row + 2, 8, contexto.get("RESPONSAVEL_DATA_MANUTENCAO_SE", ""), 9)

        bloco_label(row + 3, 1, "Assinatura Respons\u00e1vel:", 3)
        bloco_valor(row + 3, 4, contexto.get("ASSINATURA_MANUTENCAO", ""), 6)
        bloco_label(row + 3, 7, "Situa\u00e7\u00e3o:")
        bloco_valor(row + 3, 8, status, 9, center=True)

        bloco_label(row + 4, 1, "Data/Hora:", 3)
        bloco_valor(row + 4, 4, contexto.get("DATA_MANUTENCAO", ""), 6)
        bloco_label(row + 4, 7, "")
        bloco_valor(row + 4, 8, "", 9)

        titulo(row + 5, "Transfer\u00eancia de Respons\u00e1vel")
        preparar_linha(row + 6)
        preparar_linha(row + 7)
        bloco_label(row + 6, 1, "1\u00aa transfer\u00eancia", 2)
        bloco_label(row + 6, 3, "Antigo", 4)
        bloco_label(row + 6, 5, "Novo", 6)
        bloco_label(row + 6, 7, "Data/Hor\u00e1rio", 9)
        bloco_valor(row + 7, 1, "", 2)
        bloco_valor(row + 7, 3, "", 4)
        bloco_valor(row + 7, 5, "", 6)
        bloco_valor(row + 7, 7, "", 9)

        titulo(row + 8, "Libera\u00e7\u00e3o para Opera\u00e7\u00e3o")
        for offset in range(9, 13):
            preparar_linha(row + offset)

        bloco_label(row + 9, 1, "Assinatura Respons\u00e1vel:", 3)
        bloco_valor(row + 9, 4, contexto.get("ASSINATURA_OPERACAO", ""), 6)
        bloco_label(row + 9, 7, "Situa\u00e7\u00e3o:")
        bloco_valor(row + 9, 8, status, 9, center=True)

        bloco_label(row + 10, 1, "Data/Hora:", 3)
        bloco_valor(row + 10, 4, contexto.get("DATA_OPERACAO", ""), 6)
        bloco_label(row + 10, 7, "")
        bloco_valor(row + 10, 8, "", 9)

        bloco_label(row + 11, 1, "SE:")
        bloco_valor(row + 11, 2, contexto.get("RESPONSAVEL_OPERACAO_SE", ""), 3)
        bloco_label(row + 11, 4, "COT:")
        bloco_valor(row + 11, 5, contexto.get("RESPONSAVEL_OPERACAO_COT", ""), 6)
        bloco_label(row + 11, 7, "ONS:")
        bloco_valor(row + 11, 8, contexto.get("RESPONSAVEL_OPERACAO_ONS", ""), 9)

        bloco_label(row + 12, 1, "Data/Hora")
        bloco_valor(row + 12, 2, contexto.get("RESPONSAVEL_DATA_OPERACAO_SE", ""), 3)
        bloco_label(row + 12, 4, "Data/Hora")
        bloco_valor(row + 12, 5, contexto.get("RESPONSAVEL_DATA_OPERACAO_COT", ""), 6)
        bloco_label(row + 12, 7, "Data/Hora:")
        bloco_valor(row + 12, 8, contexto.get("RESPONSAVEL_DATA_OPERACAO_ONS", ""), 9)

        if liberacao.observacoes:
            preparar_linha(row + 13, 32)
            bloco_label(row + 13, 1, "Observa\u00e7\u00f5es:", 2)
            bloco_valor(row + 13, 3, limpar(liberacao.observacoes), 9)
            return row + 15

        return row + 14

    linha = 1
    if not liberacoes:
        class RegistroVazio:
            data_hora_liberacao = None
            data_hora_devolucao = None
            usuario_solicitou = None
            usuario_devolveu = None
            operador_liberou = None
            operador_recebeu_devolucao = None
            observacoes = None
            status = "PROGRAMADA"

        render_liberacao(RegistroVazio(), linha)
    else:
        for liberacao in liberacoes:
            linha = render_liberacao(liberacao, linha)

def gerar_xlsm(modelo, destino, contexto, mapeamento, liberacoes=None):

    print("Destino:", destino)
    print("Existe:", os.path.exists(destino))
    print("Tamanho:", os.path.getsize(destino) if os.path.exists(destino) else "N/A")

    if os.path.exists(destino):
        os.remove(destino)

    shutil.copy(modelo, destino)

    wb = load_workbook(destino)
    ws = wb.active

    img = Image("modelos/logo.jpg")
    img.width = 150
    img.height = 45

    ws.add_image(img, "A1")

    for campo, celula in mapeamento.items():
        valor = contexto.get(campo, "")
        set_valor_seguro(ws, celula, valor)

    adicionar_aba_liberacoes(wb, liberacoes or [], contexto)
    wb.save(destino)


def montar_contexto_si(si, ativo=None, sub=None, grupo=None):
    def primeiro(*valores):
        for valor in valores:
            if valor not in (None, ""):
                return valor
        return ""

    def sim_nao(valor):
        if valor in (None, ""):
            return ""
        return str(valor).replace("NAO", "NÃO")

    codigo_ativo = ativo.codigo_ativo if ativo else (grupo.codigo_ativo if grupo else "")
    fase = ativo.fase if ativo else ("Todas as fases" if grupo else "")
    bay = ativo.bay if ativo else (grupo.bay if grupo else "")

    localizacao = ""
    if codigo_ativo:
        localizacao = " - ".join(
            str(valor)
            for valor in (codigo_ativo, fase, bay)
            if valor not in (None, "")
        )

    tempo_retorno = limpar(getattr(si, "tempo_retorno", ""))
    disponivel = limpar(getattr(si, "disponivel", ""))
    if tempo_retorno and disponivel:
        tempo_retorno = f"{tempo_retorno} | Disponível: {disponivel}"
    elif disponivel:
        tempo_retorno = f"Disponível: {disponivel}"

    return {
        "NUM_SI": limpar(si.numero_si),
        "NUM_SGI": limpar(getattr(si, "numero_sgi", "")),
        "NUM_OS": limpar(getattr(si, "numero_os", "")),
        "NUM_APR": limpar(si.numero_apr),
        "ESPECIE": limpar(si.especie),
        "INSTALACAO": limpar(sub.nome if sub else ""),
        "LOCALIZACAO": limpar(localizacao),
        "CODIGO_ATIVO": limpar(codigo_ativo),
        "NATUREZA": limpar(getattr(si, "natureza", "")),
        "CARACTERISTICA_INTERVENCAO": limpar(getattr(si, "caracteristica_intervencao", "")),
        "TIPO": limpar(getattr(si, "tipo", "")),
        "DOCUMENTOS_REFERENCIA": limpar(getattr(si, "documentos_referencia", "")),
        "DT_INICIO_PRERIODO_TOTAL": limpar(si.data_inicio_preriodo_total),
        "DT_FIM_PRERIODO_TOTAL": limpar(si.data_fim_preriodo_total),
        "DT_INICIO_PRERIODO_MANUTENCAO": limpar(si.data_inicio_preriodo_manutencao),
        "DT_FIM_PRERIODO_MANUTENCAO": limpar(si.data_fim_preriodo_manutencao),
        "JUSTIFICATIVA": limpar(getattr(si, "justificativa", "")),
        "RESPONSAVEL": limpar(si.responsavel),
        "SUBSTITUTO": limpar(si.substituto),
        "APROVEITAMENTO": sim_nao(getattr(si, "aproveitamento", "")),
        "INCLUSAO_SERVICO": sim_nao(getattr(si, "inclusao_servico", "")),
        "ACARRETA_RISCO_PERDAS_MULTIPLAS": sim_nao(getattr(si, "acarreta_risco_perdas_multiplas", "")),
        "POSTERGACAO_TRAZ_RISCO": formatar_risco_postergacao(getattr(si, "postergacao_traz_risco", "")),
        "ORGAOS": limpar(getattr(si, "orgaos", "")),
        "TIPO_PROGRAMACAO": limpar(primeiro(getattr(si, "tipo_programacao", ""), getattr(si, "tipo_progrmacao", ""))),
        "DIAS_EXCECAO": limpar(getattr(si, "dias_excecao", "")),
        "TEMPO_RETORNO": tempo_retorno,
        "DESC_SERVICOS": limpar(si.descricao_servicos),
        "OBSERVACOES": limpar(si.observacoes),
        "CABO_ATERRAMENTO": limpar(getattr(si, "cabo_aterramento", "")),
        "RISCO_DESLIGAMENTO": sim_nao(getattr(si, "risco_desligamento", "")),
        "CONDICOES_CLIMATICAS": sim_nao(getattr(si, "condicoes_climaticas", "")),
        "EXECUCAO_PERIODO_NOTURNO": sim_nao(getattr(si, "execucao_periodo_noturno", "")),
        "RESPONSAVEL_MANUTENCAO_ONS": limpar(si.responsavel_ons_manutencao),
        "RESPONSAVEL_MANUTENCAO_COT": limpar(si.responsavel_cot_manutencao),
        "RESPONSAVEL_MANUTENCAO_SE": limpar(si.responsavel_se_manutencao),
        "RESPONSAVEL_DATA_MANUTENCAO_ONS": limpar(si.responsavel_data_ons_manutencao),
        "RESPONSAVEL_DATA_MANUTENCAO_COT": limpar(si.responsavel_data_cot_manutencao),
        "RESPONSAVEL_DATA_MANUTENCAO_SE": limpar(si.responsavel_data_se_manutencao),
        "ASSINATURA_MANUTENCAO": limpar(primeiro(si.responsavel_se_manutencao, si.responsavel_cot_manutencao, si.responsavel_ons_manutencao)),
        "STATUS_MANUTENCAO": limpar(si.status_manutencao),
        "DATA_MANUTENCAO": limpar(primeiro(si.responsavel_data_se_manutencao, si.responsavel_data_cot_manutencao, si.responsavel_data_ons_manutencao)),
        "RESPONSAVEL_OPERACAO_SE": limpar(si.responsavel_se_operacao),
        "RESPONSAVEL_OPERACAO_COT": limpar(si.responsavel_cot_operacao),
        "RESPONSAVEL_OPERACAO_ONS": limpar(si.responsavel_ons_operacao),
        "ASSINATURA_OPERACAO": limpar(primeiro(si.responsavel_se_operacao, si.responsavel_cot_operacao, si.responsavel_ons_operacao)),
        "STATUS_OPERACAO": limpar(si.status_operacao),
        "DATA_OPERACAO": limpar(primeiro(si.responsavel_data_se_operacao, si.responsavel_data_cot_operacao, si.responsavel_data_ons_operacao)),
    }


@router.post("", response_model=SIResponse)
def criar_si(dados: SICreate, db: Session = Depends(get_db)):
    garantir_colunas_si(db)
    garantir_estrutura_grupo_ativo(db)
    sincronizar_grupos_ativos(db)

    data = dados.dict()
    validar_selecao_ativo(db, data.get("id_subestacao"), data.get("id_funcao_operacao"), data.get("id_grupo_ativo"), data.get("escopo_ativo"), data.get("id_ativo"))
    if data.get("id_ativo"):
        ativo = db.query(Ativo).filter(Ativo.id_ativo == data["id_ativo"]).first()
        if ativo:
            data["especie"] = especie_documento_por_ativo(ativo) or data.get("especie")

    data["prioridade"] = normalizar_prioridade_operacao(data.get("prioridade"))

    if not data.get("numero_si"):
        data["numero_si"] = gerar_numero_si(db, sigla_por_subestacao(data.get("id_subestacao")))

    nova_si = solicitacao_intervencao(**data)
    db.add(nova_si)
    db.commit()
    db.refresh(nova_si)
    return nova_si


@router.get("", response_model=list[SIResponse])
def listar_si(db: Session = Depends(get_db)):
    garantir_colunas_si(db)
    return (
        db.query(solicitacao_intervencao)
        .options(selectinload(solicitacao_intervencao.ativo))
        .order_by(solicitacao_intervencao.id_si.desc())
        .all()
    )


@router.get("/paginado", response_model=SIPaginadaResponse)
def listar_si_paginado(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    search: str | None = None,
    status: str | None = None,
    id_subestacao: int | None = None,
    db: Session = Depends(get_db),
):
    garantir_colunas_si(db)
    query = db.query(solicitacao_intervencao).outerjoin(
        Ativo, solicitacao_intervencao.id_ativo == Ativo.id_ativo
    )

    if search and search.strip():
        termo = f"%{search.strip()}%"
        query = query.filter(or_(
            solicitacao_intervencao.numero_si.ilike(termo),
            solicitacao_intervencao.descricao_servicos.ilike(termo),
            Ativo.codigo_ativo.ilike(termo),
        ))
    if status and status != "all":
        query = query.filter(solicitacao_intervencao.status_manutencao == status)
    if id_subestacao:
        query = query.filter(solicitacao_intervencao.id_subestacao == id_subestacao)

    total = query.count()
    items = (
        query.options(selectinload(solicitacao_intervencao.ativo))
        .order_by(solicitacao_intervencao.id_si.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
    }


@router.get("/{id_si}", response_model=SIResponse)
def buscar_si(id_si: int, db: Session = Depends(get_db)):
    garantir_colunas_si(db)
    return obter_si_ou_404(db, id_si)


@router.get("/{id_si}/liberacoes", response_model=list[SILiberacaoResponse])
def listar_liberacoes_si(id_si: int, db: Session = Depends(get_db)):
    garantir_tabela_liberacoes_si(db)
    obter_si_ou_404(db, id_si)
    return (
        db.query(SILiberacao)
        .filter(SILiberacao.id_si == id_si)
        .order_by(SILiberacao.data_hora_liberacao.desc(), SILiberacao.id_liberacao.desc())
        .all()
    )


@router.post("/{id_si}/liberacoes/manutencao", response_model=SILiberacaoResponse)
def liberar_si_para_manutencao(
    id_si: int,
    dados: SILiberacaoManutencaoCreate,
    db: Session = Depends(get_db),
    usuario=Depends(get_current_user),
):
    garantir_tabela_liberacoes_si(db)
    obter_si_ou_404(db, id_si)

    liberacao_aberta = (
        db.query(SILiberacao)
        .filter(SILiberacao.id_si == id_si, SILiberacao.status == "EM_EXECUCAO")
        .first()
    )
    if liberacao_aberta:
        raise HTTPException(400, "J? existe uma libera??o em execu??o para esta SI")

    agora = datetime.utcnow()
    liberacao = SILiberacao(
        id_si=id_si,
        data_utilizacao=dados.data_utilizacao or date.today(),
        data_hora_liberacao=dados.data_hora_liberacao or agora,
        usuario_solicitou_id=getattr(usuario, "id", None),
        usuario_solicitou=nome_usuario(usuario),
        operador_liberou=dados.operador_liberou,
        observacoes=dados.observacoes,
        status="EM_EXECUCAO",
        criado_em=agora,
        atualizado_em=agora,
    )

    db.add(liberacao)
    db.commit()
    db.refresh(liberacao)
    return liberacao


@router.put("/{id_si}/liberacoes/{id_liberacao}/operacao", response_model=SILiberacaoResponse)
def liberar_si_para_operacao(
    id_si: int,
    id_liberacao: int,
    dados: SILiberacaoOperacaoUpdate,
    db: Session = Depends(get_db),
    usuario=Depends(get_current_user),
):
    garantir_tabela_liberacoes_si(db)
    obter_si_ou_404(db, id_si)

    liberacao = (
        db.query(SILiberacao)
        .filter(SILiberacao.id_si == id_si, SILiberacao.id_liberacao == id_liberacao)
        .first()
    )
    if not liberacao:
        raise HTTPException(404, "Libera??o n?o encontrada")
    if liberacao.status != "EM_EXECUCAO":
        raise HTTPException(400, "A libera??o n?o est? em execu??o")

    agora = datetime.utcnow()
    liberacao.data_hora_devolucao = dados.data_hora_devolucao or agora
    liberacao.usuario_devolveu_id = getattr(usuario, "id", None)
    liberacao.usuario_devolveu = nome_usuario(usuario)
    liberacao.operador_recebeu_devolucao = dados.operador_recebeu_devolucao
    if dados.observacoes is not None:
        liberacao.observacoes = dados.observacoes
    liberacao.status = "ENCERRADA"
    liberacao.atualizado_em = agora

    db.commit()
    db.refresh(liberacao)
    return liberacao


@router.put("/{id_si}/liberacoes/{id_liberacao}/cancelar", response_model=SILiberacaoResponse)
def cancelar_liberacao_si(
    id_si: int,
    id_liberacao: int,
    dados: SILiberacaoCancelarUpdate,
    db: Session = Depends(get_db),
    usuario=Depends(get_current_user),
):
    garantir_tabela_liberacoes_si(db)
    obter_si_ou_404(db, id_si)

    liberacao = (
        db.query(SILiberacao)
        .filter(SILiberacao.id_si == id_si, SILiberacao.id_liberacao == id_liberacao)
        .first()
    )
    if not liberacao:
        raise HTTPException(404, "Libera??o n?o encontrada")
    if liberacao.status != "EM_EXECUCAO":
        raise HTTPException(400, "Apenas Liberacoes em execu??o podem ser canceladas")

    agora = datetime.utcnow()
    liberacao.status = "CANCELADA"
    liberacao.data_hora_devolucao = agora
    liberacao.usuario_devolveu_id = getattr(usuario, "id", None)
    liberacao.usuario_devolveu = nome_usuario(usuario)
    if dados.observacoes is not None:
        liberacao.observacoes = dados.observacoes
    liberacao.atualizado_em = agora

    db.commit()
    db.refresh(liberacao)
    return liberacao


@router.put("/{id_si}", response_model=SIResponse)
def editar_si(id_si: int, dados: SIUpdate, db: Session = Depends(get_db), usuario=Depends(get_current_user)):
    garantir_colunas_si(db)
    si = obter_si_ou_404(db, id_si)

    for campo, valor in dados.dict(exclude_unset=True).items():
        if campo == "emissor":
            continue
        setattr(si, campo, valor)

    si.editado_por = getattr(usuario, "nome", None) or getattr(usuario, "email", None)

    if si.id_ativo:
        ativo = db.query(Ativo).filter(Ativo.id_ativo == si.id_ativo).first()
        if ativo:
            si.especie = especie_documento_por_ativo(ativo) or si.especie

    si.prioridade = normalizar_prioridade_operacao(si.prioridade)
    db.commit()
    db.refresh(si)
    return si


@router.delete("/{id_si}")
def deletar_si(id_si: int, db: Session = Depends(get_db), _usuario=Depends(require_roles("admin"))):
    si = db.query(solicitacao_intervencao).filter(solicitacao_intervencao.id_si == id_si).first()
    if not si:
        raise HTTPException(404, "SI n?o encontrada")

    db.delete(si)
    db.commit()
    return {"message": "SI deletada com sucesso"}


@router.get("/{id_si}/download")
def download_si(id_si: int, db: Session = Depends(get_db)):
    garantir_colunas_si(db)
    garantir_tabela_liberacoes_si(db)
    si = obter_si_ou_404(db, id_si)

    ativo = db.query(Ativo).filter(Ativo.id_ativo == si.id_ativo).first() if si.id_ativo else None
    grupo = (
        db.query(GrupoAtivo).filter(GrupoAtivo.id_grupo_ativo == si.id_grupo_ativo).first()
        if si.id_grupo_ativo
        else None
    )
    sub = db.query(Subestacao).filter(Subestacao.id_subestacao == si.id_subestacao).first() if si.id_subestacao else None
    liberacoes = (
        db.query(SILiberacao)
        .filter(SILiberacao.id_si == id_si)
        .order_by(SILiberacao.data_hora_liberacao.asc(), SILiberacao.id_liberacao.asc())
        .all()
    )

    contexto = montar_contexto_si(si, ativo, sub, grupo)
    pasta_saida = "saida"
    os.makedirs(pasta_saida, exist_ok=True)

    nome_arquivo = nome_arquivo_seguro(f"{si.numero_si}.xlsx")
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    gerar_xlsm(
        modelo="modelos/MODELO_SI.xlsx",
        destino=caminho_saida,
        contexto=contexto,
        mapeamento=MAPEAMENTO_CELULAS,
        liberacoes=liberacoes,
    )

    return FileResponse(path=caminho_saida, filename=nome_arquivo, media_type="application/octet-stream")


