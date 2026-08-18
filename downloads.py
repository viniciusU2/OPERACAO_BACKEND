import os
import re
import tempfile
from datetime import date, datetime
from decimal import Decimal
from enum import Enum

from fastapi import APIRouter, Depends, Query
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text

from database import get_db
from models.Ativo import Ativo
from models.OS_models import OrdemServico
from models.SI_models import solicitacao_intervencao
from models.SS_models import SolicitacaoServico
from models.plano_manutencao_models import Inspecao, ResultadoItemInspecao


router = APIRouter(prefix="/downloads", tags=["Downloads"])


def nome_arquivo_seguro(texto: str):
    return re.sub(r"[^A-Za-z0-9_.-]", "_", texto)


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, Enum):
        return valor.value
    return valor


def valor_campo(registro, campo: str):
    valor = registro
    for parte in campo.split("."):
        valor = getattr(valor, parte, None)
        if valor is None:
            return ""
    return valor


def aplicar_estilo(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def garantir_colunas_si(db: Session):
    colunas = {
        "natureza": "VARCHAR(255) NULL",
        "caracteristica_intervencao": "VARCHAR(100) NULL",
    }

    for coluna, definicao in colunas.items():
        existe = db.execute(
            text("SHOW COLUMNS FROM solicitacao_intervencao LIKE :coluna"),
            {"coluna": coluna},
        ).first()
        if not existe:
            db.execute(text(f"ALTER TABLE solicitacao_intervencao ADD COLUMN {coluna} {definicao}"))

    db.commit()


def adicionar_aba(wb, titulo: str, colunas: list[tuple[str, str]], registros):
    ws = wb.create_sheet(titulo)
    
    ws.append([label for label, _ in colunas])

    for registro in registros:
        ws.append([
            limpar(valor_campo(registro, campo))
            for _, campo in colunas
        ])

    aplicar_estilo(ws)


def filtrar_por_intervalo(query, modelo, campo_inicio: str | None, data_inicio, data_fim):
    if not campo_inicio:
        return query

    coluna = getattr(modelo, campo_inicio, None)
    if coluna is None:
        return query

    if data_inicio:
        query = query.filter(coluna >= data_inicio)
    if data_fim:
        query = query.filter(coluna <= data_fim)

    return query


def aplicar_filtros_operacionais(
    query,
    modelo,
    status: str | None = None,
    id_subestacao: int | None = None,
    data_inicio=None,
    data_fim=None,
    campo_data: str | None = None,
):
    if status and status != "all":
        if modelo is solicitacao_intervencao:
            query = query.filter(
                (modelo.status_manutencao == status) | (modelo.status_operacao == status)
            )
        elif hasattr(modelo, "status"):
            query = query.filter(modelo.status == status)

    if id_subestacao:
        query = query.filter(modelo.id_subestacao == id_subestacao)

    return filtrar_por_intervalo(query, modelo, campo_data, data_inicio, data_fim)


OS_COLUNAS = [
    ("ID", "id_os"),
    ("Numero OS", "numero_os"),
    ("Numero SI", "numero_si"),
    ("Subestacao", "id_subestacao"),
    ("Ativo", "id_ativo"),
    # A OS de todas as fases referencia o grupo, sem um ativo individual.
    # A propriedade da OS resolve o codigo tanto do ativo quanto do grupo.
    ("Codigo Ativo", "codigo_ativo"),
    ("Especie", "especie"),
    ("APR", "numero_apr"),
    ("Instalacao", "instalacao"),
    ("Localizacao", "localizacao"),
    ("Complemento", "complemento"),
    ("Origem", "origens"),
    ("Defeito", "defeito"),
    ("Esquema Servicos", "esquema_servicos"),
    ("Prioridade", "prioridade"),
    ("Responsavel", "responsavel"),
    ("Resp. Manutencao", "responsavel_manutencao"),
    ("Resp. Operacao", "responsavel_operacao"),
    ("Substituto", "substituto"),
    ("Inicio Programado", "data_inicio_programado"),
    ("Fim Programado", "data_fim_programado"),
    ("Descricao Servicos", "descricao_servicos"),
    ("Observacoes", "observacoes"),
    ("Causa Primaria", "causa_primaria"),
    ("Causa Secundaria", "causa_secundaria"),
    ("Emissor", "emissor"),
    ("Abertura SS", "data_abertura_ss"),
    ("Inicio Execucao", "data_inicio_execucao"),
    ("Fim Execucao", "data_fim_execucao"),
    ("Centro Custos", "centro_custos"),
    ("Status", "status"),
    ("Criado em", "criado_em"),
]


SI_COLUNAS = [
    ("ID", "id_si"),
    ("Numero SI", "numero_si"),
    ("Numero SGI", "numero_sgi"),
    ("Subestacao", "id_subestacao"),
    ("Ativo", "id_ativo"),
    ("Especie", "especie"),
    ("APR", "numero_apr"),
    ("Natureza", "natureza"),
    ("Caracterizacao Intervencao", "caracteristica_intervencao"),
    ("Tipo", "tipo"),
    ("Documentos Referencia", "documentos_referencia"),
    ("Inicio Periodo Total", "data_inicio_preriodo_total"),
    ("Fim Periodo Total", "data_fim_preriodo_total"),
    ("Inicio Manutencao", "data_inicio_preriodo_manutencao"),
    ("Fim Manutencao", "data_fim_preriodo_manutencao"),
    ("Justificativa", "justificativa"),
    ("Responsavel", "responsavel"),
    ("Substituto", "substituto"),
    ("Aproveitamento", "aproveitamento"),
    ("Inclusao Servico", "inclusao_servico"),
    ("Orgaos", "orgaos"),
    ("Tipo Programacao", "tipo_programacao"),
    ("Dias Excecao", "dias_excecao"),
    ("Tempo Retorno", "tempo_retorno"),
    ("Disponivel", "disponivel"),
    ("Descricao Servicos", "descricao_servicos"),
    ("Observacoes", "observacoes"),
    ("Cabo Aterramento", "cabo_aterramento"),
    ("Risco Desligamento", "risco_desligamento"),
    ("Condicoes Climaticas", "condicoes_climaticas"),
    ("Periodo Noturno", "execucao_periodo_noturno"),
    ("Resp. ONS Manutencao", "responsavel_ons_manutencao"),
    ("Resp. COT Manutencao", "responsavel_cot_manutencao"),
    ("Resp. SE Manutencao", "responsavel_se_manutencao"),
    ("Data ONS Manutencao", "responsavel_data_ons_manutencao"),
    ("Data COT Manutencao", "responsavel_data_cot_manutencao"),
    ("Data SE Manutencao", "responsavel_data_se_manutencao"),
    ("Status Manutencao", "status_manutencao"),
    ("Resp. ONS Operacao", "responsavel_ons_operacao"),
    ("Resp. COT Operacao", "responsavel_cot_operacao"),
    ("Resp. SE Operacao", "responsavel_se_operacao"),
    ("Data ONS Operacao", "responsavel_data_ons_operacao"),
    ("Data COT Operacao", "responsavel_data_cot_operacao"),
    ("Data SE Operacao", "responsavel_data_se_operacao"),
    ("Status Operacao", "status_operacao"),
    ("Emissor", "emissor"),
    ("Criado em", "criado_em"),
]


SS_COLUNAS = [
    ("ID", "id"),
    ("Numero SS", "numero_ss"),
    ("Numero OS", "numero_os"),
    ("Data Solicitacao", "data_hora_solicitacao"),
    ("Data Abertura", "data_hora_abertura"),
    ("Data Limite", "data_hora_limite"),
    ("Solicitante", "solicitante"),
    ("Matricula", "matricula"),
    ("Funcao", "funcao"),
    ("Telefone", "telefone"),
    ("Email", "email"),
    ("Orgao", "orgao"),
    ("Instalacao", "instalacao"),
    ("Localizacao", "localizacao"),
    ("Complemento", "complemento"),
    ("Ativo", "id_ativo"),
    ("Codigo Ativo", "codigo_ativo"),
    ("Tipo Ativo", "tipo_ativo_nome"),
    ("Fase", "fase_ativo"),
    ("Esquema Servico", "esquema_servico"),
    ("Centro Custo", "centro_custo"),
    ("Causa", "causa"),
    ("Causa Secundaria", "causa_secundaria"),
    ("Equipe", "equipe"),
    ("Descricao Problema", "descricao_problema"),
    ("Problemas Tipicos", "problemas_titulos"),
    ("Criticidades Identificadas", "problemas_criticidades"),
    ("Observacoes Problemas", "problemas_observacoes"),
    ("Problemas Confirmados", "problemas_confirmados"),
    ("Prioridade", "prioridade"),
    ("Status", "status"),
]


ATIVO_COLUNAS = [
    ("Subestação", "subestacao.nome"),
    ("Tipo de ativo", "tipo_ativo.nome"),
    ("Código Função de Transmissão", "funcao_operacao.codigo"),
    ("Descrição Função de Transmissão", "funcao_operacao.descricao"),
    ("Codigo Ativo", "codigo_ativo"),
    ("Fabricante", "fabricante"),
    ("Modelo", "modelo"),
    ("Especie", "especie"),
    ("Numero Serie", "numero_serie"),
    ("Tensao Nominal kV", "tensao_nominal_kv"),
    ("Data Instalacao", "data_instalacao"),
    ("Status", "status"),
    ("Bay", "bay"),
    ("Fase", "fase"),
]


INSPECAO_COLUNAS = [
    ("ID", "id_inspecao"),
    ("Instalacao", "ativo.subestacao.nome"),
    ("ID Ativo", "id_ativo"),
    ("Codigo Ativo", "codigo_ativo"),
    ("Tipo de Ativo", "ativo.tipo_ativo.nome"),
    ("Fase", "fase_ativo"),
    ("Bay", "ativo.bay"),
    ("Numero OS", "ordem_servico.numero_os"),
    ("Data Inspecao", "data_inspecao"),
    ("Proxima Inspecao", "data_proxima_inspecao"),
    ("Periodicidade", "periodicidade"),
    ("Responsavel", "responsavel"),
    ("Resultado Geral", "status_geral"),
    ("Observacao Geral", "observacao_geral"),
    ("Ficha de Inspecao", "ficha_inspecao_url"),
]


RESULTADO_INSPECAO_COLUNAS = [
    "ID Inspecao",
    "Instalacao",
    "ID Ativo",
    "Codigo Ativo",
    "Tipo de Ativo",
    "Fase",
    "Bay",
    "Numero OS",
    "Data Inspecao",
    "Periodicidade",
    "Responsavel",
    "Resultado Geral",
    "Item Inspecionado",
    "Unidade",
    "Valor de Referencia",
    "Tolerancia",
    "Valor Medido",
    "Resultado do Item",
    "Observacao do Item",
    "Foto",
]


def adicionar_aba_resultados_inspecao(wb, inspecoes):
    ws = wb.create_sheet("Resultados detalhados")
    ws.append(RESULTADO_INSPECAO_COLUNAS)

    for inspecao in inspecoes:
        for resultado in inspecao.resultados:
            ws.append([
                limpar(inspecao.id_inspecao),
                limpar(valor_campo(inspecao, "ativo.subestacao.nome")),
                limpar(inspecao.id_ativo),
                limpar(valor_campo(inspecao, "ativo.codigo_ativo")),
                limpar(valor_campo(inspecao, "ativo.tipo_ativo.nome")),
                limpar(valor_campo(inspecao, "ativo.fase")),
                limpar(valor_campo(inspecao, "ativo.bay")),
                limpar(valor_campo(inspecao, "ordem_servico.numero_os")),
                limpar(inspecao.data_inspecao),
                limpar(inspecao.periodicidade),
                limpar(inspecao.responsavel),
                limpar(inspecao.status_geral),
                limpar(resultado.nome_item),
                limpar(resultado.unidade),
                limpar(resultado.valor_referencia),
                limpar(resultado.tolerancia),
                limpar(resultado.valor_medido),
                limpar(resultado.status_item),
                limpar(resultado.observacao_item),
                limpar(resultado.foto),
            ])

    aplicar_estilo(ws)


def adicionar_aba_indicadores_inspecao(wb, inspecoes):
    ws = wb.create_sheet("Indicadores", 0)
    ws["A1"] = "INDICADORES DAS INSPECOES"
    ws["A1"].font = Font(size=16, bold=True, color="1F2937")

    contagem_itens = {"OK": 0, "NOK": 0, "NA": 0}
    contagem_geral = {"OK": 0, "NOK": 0, "NA": 0}

    for inspecao in inspecoes:
        status_geral = limpar(inspecao.status_geral)
        contagem_geral[status_geral] = contagem_geral.get(status_geral, 0) + 1
        for resultado in inspecao.resultados:
            status_item = limpar(resultado.status_item)
            contagem_itens[status_item] = contagem_itens.get(status_item, 0) + 1

    ws.append([])
    ws.append(["Resultado dos itens", "Quantidade"])
    for status in ("OK", "NOK", "NA"):
        ws.append([status, contagem_itens.get(status, 0)])

    ws["D3"] = "Resultado geral"
    ws["E3"] = "Quantidade"
    for linha, status in enumerate(("OK", "NOK", "NA"), start=4):
        ws.cell(linha, 4, status)
        ws.cell(linha, 5, contagem_geral.get(status, 0))

    for cell in list(ws[3])[0:2] + list(ws[3])[3:5]:
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True)

    grafico_itens = BarChart()
    grafico_itens.type = "col"
    grafico_itens.style = 10
    grafico_itens.title = "Resultados por item inspecionado"
    grafico_itens.y_axis.title = "Quantidade"
    grafico_itens.x_axis.title = "Resultado"
    grafico_itens.height = 8
    grafico_itens.width = 13
    grafico_itens.add_data(Reference(ws, min_col=2, min_row=3, max_row=6), titles_from_data=True)
    grafico_itens.set_categories(Reference(ws, min_col=1, min_row=4, max_row=6))
    ws.add_chart(grafico_itens, "A9")

    grafico_geral = BarChart()
    grafico_geral.type = "col"
    grafico_geral.style = 11
    grafico_geral.title = "Resultado geral das inspecoes"
    grafico_geral.y_axis.title = "Quantidade"
    grafico_geral.x_axis.title = "Resultado"
    grafico_geral.height = 8
    grafico_geral.width = 13
    grafico_geral.add_data(Reference(ws, min_col=5, min_row=3, max_row=6), titles_from_data=True)
    grafico_geral.set_categories(Reference(ws, min_col=4, min_row=4, max_row=6))
    ws.add_chart(grafico_geral, "H9")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14


def salvar_workbook(wb, nome_base: str):
    pasta_saida = tempfile.mkdtemp(prefix="downloads_")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = nome_arquivo_seguro(f"{nome_base}_{timestamp}.xlsx")
    caminho = os.path.join(pasta_saida, nome_arquivo)

    wb.save(caminho)
    return caminho, nome_arquivo


@router.get("/operacionais")
def baixar_operacionais(
    documento: str = Query(default="all"),
    status: str | None = Query(default=None),
    id_subestacao: int | None = Query(default=None),
    data_inicio: datetime | None = Query(default=None),
    data_fim: datetime | None = Query(default=None),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    wb.remove(wb.active)

    documento = (documento or "all").lower()

    if documento in ("all", "os"):
        query = aplicar_filtros_operacionais(
            db.query(OrdemServico),
            OrdemServico,
            status=status,
            id_subestacao=id_subestacao,
            data_inicio=data_inicio,
            data_fim=data_fim,
            campo_data="data_inicio_programado",
        )
        adicionar_aba(wb, "OS", OS_COLUNAS, query.all())

    if documento in ("all", "si"):
        garantir_colunas_si(db)
        query = aplicar_filtros_operacionais(
            db.query(solicitacao_intervencao),
            solicitacao_intervencao,
            status=status,
            id_subestacao=id_subestacao,
            data_inicio=data_inicio,
            data_fim=data_fim,
            campo_data="data_inicio_preriodo_total",
        )
        adicionar_aba(wb, "SI", SI_COLUNAS, query.all())

    if documento in ("all", "ss"):
        query = aplicar_filtros_operacionais(
            db.query(SolicitacaoServico),
            SolicitacaoServico,
            status=status,
            id_subestacao=None,
            data_inicio=data_inicio,
            data_fim=data_fim,
            campo_data="data_hora_solicitacao",
        )
        if id_subestacao:
            query = query.join(Ativo, Ativo.id_ativo == SolicitacaoServico.id_ativo).filter(
                Ativo.id_subestacao == id_subestacao
            )
        adicionar_aba(wb, "SS", SS_COLUNAS, query.all())

    caminho, nome_arquivo = salvar_workbook(wb, f"operacionais_{documento}")

    return FileResponse(
        path=caminho,
        filename=nome_arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/ativos")
def baixar_ativos(
    status: str | None = Query(default=None),
    id_subestacao: int | None = Query(default=None),
    id_tipo_ativo: int | None = Query(default=None),
    data_inicio: datetime | None = Query(default=None),
    data_fim: datetime | None = Query(default=None),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    query = db.query(Ativo).options(
        joinedload(Ativo.subestacao),
        joinedload(Ativo.tipo_ativo),
        joinedload(Ativo.funcao_operacao),
    )
    if status and status != "all":
        query = query.filter(Ativo.status == status)
    if id_subestacao:
        query = query.filter(Ativo.id_subestacao == id_subestacao)
    if id_tipo_ativo:
        query = query.filter(Ativo.id_tipo_ativo == id_tipo_ativo)
    query = filtrar_por_intervalo(query, Ativo, "data_instalacao", data_inicio, data_fim)

    ws.append([label for label, _ in ATIVO_COLUNAS])
    for ativo in query.all():
        ws.append([
              limpar(valor_campo(ativo, campo))
            for _, campo in ATIVO_COLUNAS
        ])

    aplicar_estilo(ws)
    caminho, nome_arquivo = salvar_workbook(wb, "ativos")

    return FileResponse(
        path=caminho,
        filename=nome_arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/inspecoes")
def baixar_inspecoes(
    status: str | None = Query(default=None),
    id_subestacao: int | None = Query(default=None),
    id_tipo_ativo: int | None = Query(default=None),
    data_inicio: datetime | None = Query(default=None),
    data_fim: datetime | None = Query(default=None),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    wb.remove(wb.active)

    query = db.query(Inspecao).join(Ativo, Inspecao.id_ativo == Ativo.id_ativo).options(
        joinedload(Inspecao.ativo).joinedload(Ativo.subestacao),
        joinedload(Inspecao.ativo).joinedload(Ativo.tipo_ativo),
        joinedload(Inspecao.ordem_servico),
        joinedload(Inspecao.resultados).joinedload(ResultadoItemInspecao.plano_item),
    )

    if status and status != "all":
        query = query.filter(Inspecao.status_geral == status)
    if id_subestacao:
        query = query.filter(Ativo.id_subestacao == id_subestacao)
    if id_tipo_ativo:
        query = query.filter(Ativo.id_tipo_ativo == id_tipo_ativo)

    query = filtrar_por_intervalo(
        query,
        Inspecao,
        "data_inspecao",
        data_inicio,
        data_fim,
    )

    inspecoes = query.order_by(Inspecao.data_inspecao.desc()).all()

    adicionar_aba(
        wb,
        "Inspecoes",
        INSPECAO_COLUNAS,
        inspecoes,
    )
    adicionar_aba_resultados_inspecao(wb, inspecoes)
    adicionar_aba_indicadores_inspecao(wb, inspecoes)
    caminho, nome_arquivo = salvar_workbook(wb, "inspecoes")

    return FileResponse(
        path=caminho,
        filename=nome_arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
