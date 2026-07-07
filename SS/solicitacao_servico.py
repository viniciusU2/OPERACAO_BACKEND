import os
import re
import shutil
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries
from sqlalchemy.orm import Session

from database import get_db
from auth.dependencies import require_roles
from models.Ativo import Ativo
from models.OS_models import OrdemServico
from models.SS_models import SolicitacaoServico
from models.instalacao_models import Subestacao
from SS.schemas import (
    SolicitacaoServicoCreate,
    SolicitacaoServicoResponse,
    SolicitacaoServicoUpdate,
)
from utils.documentos_operacao import especie_documento_por_ativo


router = APIRouter(prefix="/ss", tags=["Solicitacao de Servico"])

SUBESTACOES_SIGLAS = ["BJD", "GOR", "JAB"]


MAPEAMENTO_CELULAS = {
    "NUM_SS": "H1",
    "DT_SOLICITACAO": "A4",
    "DT_ABERTURA": "D4",
    "DT_LIMITE": "G4",
    "SOLICITANTE": "A6",
    "MATRICULA": "D6",
    "FUNCAO": "G6",
    "TELEFONE": "A8",
    "EMAIL": "D8",
    "ORGAO": "G8",
    "INSTALACAO": "A10",
    "LOCALIZACAO": "D10",
    "COMPLEMENTO": "G10",
    "CODIGO_ATIVO": "A12",
    "ESQUEMA_SERVICO": "D12",
    "PRIORIDADE": "G12",
    "DESC_PROBLEMA": "A14",
    "CAUSA": "A18",
    "CAUSA_SECUNDARIA": "D18",
    "EQUIPE": "G18",
    "CENTRO_CUSTO": "A20",
    "STATUS": "G20",
}


def nome_arquivo_seguro(texto: str):
    if not texto:
        return "sem_nome"
    return re.sub(r"[^A-Za-z0-9_.-]", "_", texto)


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    return str(valor)


def set_valor_seguro(ws, celula, valor):
    for merged_range in ws.merged_cells.ranges:
        if celula in merged_range:
            min_col, min_row, _, _ = range_boundaries(str(merged_range))
            ws.cell(row=min_row, column=min_col).value = valor
            return
    ws[celula] = valor


def sigla_por_ativo(db: Session, id_ativo: int | None):
    if not id_ativo:
        return "GERAL"

    ativo = db.query(Ativo).filter(Ativo.id_ativo == id_ativo).first()
    if not ativo or not ativo.id_subestacao:
        return "GERAL"

    try:
        return SUBESTACOES_SIGLAS[ativo.id_subestacao - 1]
    except IndexError:
        return "GERAL"


def sigla_por_subestacao(id_subestacao: int | None):
    if not id_subestacao:
        return "GERAL"

    try:
        return SUBESTACOES_SIGLAS[id_subestacao - 1]
    except IndexError:
        return "GERAL"


def gerar_numero_ss(db: Session, sigla: str):
    ano_atual = datetime.now().year
    registros = (
        db.query(SolicitacaoServico.numero_ss)
        .filter(SolicitacaoServico.numero_ss.like(f"SS-{sigla}-%-{ano_atual}"))
        .all()
    )

    numeros = []
    for (numero_ss,) in registros:
        match = re.search(rf"SS-{sigla}-(\d+)-{ano_atual}", numero_ss or "")
        if match:
            numeros.append(int(match.group(1)))

    proximo = max(numeros) + 1 if numeros else 1
    return f"SS-{sigla}-{str(proximo).zfill(4)}-{ano_atual}"


def gerar_numero_os_atendimento_ss(db: Session, sigla: str, codigo_ativo: str | None):
    ano_atual = datetime.now().year
    registros = (
        db.query(OrdemServico.numero_os)
        .filter(OrdemServico.numero_os.like(f"OS-{sigla}-%-{ano_atual}%"))
        .all()
    )

    numeros = []
    for (numero_os,) in registros:
        match = re.search(rf"OS-{sigla}-(\d+)-{ano_atual}", numero_os or "")
        if match:
            numeros.append(int(match.group(1)))

    proximo = max(numeros) + 1 if numeros else 1
    numero_formatado = str(proximo).zfill(4)
    numero_os = f"OS-{sigla}-{numero_formatado}-{ano_atual}"
    numero_apr = f"APR-{sigla}-{numero_formatado}-{ano_atual}"

    if codigo_ativo:
        codigo_seguro = re.sub(r"[^A-Za-z0-9\-]", "", codigo_ativo)
        numero_os = f"{numero_os}-{codigo_seguro}"

    return numero_os, numero_apr


def gerar_xlsx(modelo, destino, contexto, mapeamento):
    if os.path.exists(destino):
        os.remove(destino)

    shutil.copy(modelo, destino)

    wb = load_workbook(destino)
    ws = wb.active

    logo_path = "modelos/logo.jpg"
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 150
        img.height = 45
        ws.add_image(img, "A1")

    for campo, celula in mapeamento.items():
        set_valor_seguro(ws, celula, contexto.get(campo, ""))

    wb.save(destino)


def montar_contexto_ss(ss, ativo=None):
    return {
        "NUM_SS": limpar(ss.numero_ss),
        "DT_SOLICITACAO": limpar(ss.data_hora_solicitacao),
        "DT_ABERTURA": limpar(ss.data_hora_abertura),
        "DT_LIMITE": limpar(ss.data_hora_limite),
        "SOLICITANTE": limpar(ss.solicitante),
        "MATRICULA": limpar(ss.matricula),
        "FUNCAO": limpar(ss.funcao),
        "TELEFONE": limpar(ss.telefone),
        "EMAIL": limpar(ss.email),
        "ORGAO": limpar(ss.orgao),
        "INSTALACAO": limpar(ss.instalacao),
        "LOCALIZACAO": limpar(ss.localizacao),
        "COMPLEMENTO": limpar(ss.complemento),
        "CODIGO_ATIVO": limpar(ativo.codigo_ativo if ativo else ""),
        "ESQUEMA_SERVICO": limpar(ss.esquema_servico),
        "CENTRO_CUSTO": limpar(ss.centro_custo),
        "CAUSA": limpar(ss.causa),
        "CAUSA_SECUNDARIA": limpar(ss.causa_secundaria),
        "EQUIPE": limpar(ss.equipe),
        "DESC_PROBLEMA": limpar(ss.descricao_problema),
        "PRIORIDADE": limpar(ss.prioridade),
        "STATUS": limpar(ss.status),
    }


@router.post("", response_model=SolicitacaoServicoResponse)
def criar_ss(ss: SolicitacaoServicoCreate, db: Session = Depends(get_db)):
    data = ss.model_dump()
    id_subestacao = data.pop("id_subestacao", None)
    if not data.get("numero_ss"):
        sigla = sigla_por_subestacao(id_subestacao)
        if sigla == "GERAL":
            sigla = sigla_por_ativo(db, data.get("id_ativo"))

        data["numero_ss"] = gerar_numero_ss(
            db,
            sigla,
        )

    nova_ss = SolicitacaoServico(**data)

    db.add(nova_ss)
    db.commit()
    db.refresh(nova_ss)

    return nova_ss


@router.get("", response_model=list[SolicitacaoServicoResponse])
def listar_ss(db: Session = Depends(get_db)):
    return db.query(SolicitacaoServico).all()


@router.get("/{id_ss}", response_model=SolicitacaoServicoResponse)
def buscar_ss(id_ss: int, db: Session = Depends(get_db)):
    ss = db.query(SolicitacaoServico).filter(
        SolicitacaoServico.id == id_ss
    ).first()

    if not ss:
        raise HTTPException(404, "SS nao encontrada")

    return ss


@router.put("/{id_ss}", response_model=SolicitacaoServicoResponse)
def editar_ss(
    id_ss: int,
    dados: SolicitacaoServicoUpdate,
    db: Session = Depends(get_db),
):
    ss = db.query(SolicitacaoServico).filter(
        SolicitacaoServico.id == id_ss
    ).first()

    if not ss:
        raise HTTPException(404, "SS nao encontrada")

    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(ss, campo, valor)

    db.commit()
    db.refresh(ss)

    return ss


@router.delete("/{id_ss}")
def deletar_ss(
    id_ss: int,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin")),
):
    ss = db.query(SolicitacaoServico).filter(
        SolicitacaoServico.id == id_ss
    ).first()

    if not ss:
        raise HTTPException(404, "SS nao encontrada")

    db.delete(ss)
    db.commit()

    return {"message": "SS deletada com sucesso"}


@router.post("/{id_ss}/atender")
def atender_ss(
    id_ss: int,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    ss = db.query(SolicitacaoServico).filter(
        SolicitacaoServico.id == id_ss
    ).first()

    if not ss:
        raise HTTPException(404, "SS nao encontrada")

    os_existente = db.query(OrdemServico).filter(
        OrdemServico.numero_ss == ss.numero_ss
    ).first()

    if os_existente:
        ss.numero_os = os_existente.numero_os
        db.commit()
        db.refresh(ss)
        return {
            "message": "SS ja possui OS vinculada",
            "ss": ss,
            "os": os_existente,
        }

    ativo = db.query(Ativo).filter(
        Ativo.id_ativo == ss.id_ativo
    ).first() if ss.id_ativo else None

    if not ativo or not ativo.id_subestacao:
        raise HTTPException(
            status_code=400,
            detail="SS precisa estar vinculada a um ativo com subestacao para gerar OS",
        )

    subestacao = db.query(Subestacao).filter(
        Subestacao.id_subestacao == ativo.id_subestacao
    ).first()

    if not subestacao:
        raise HTTPException(400, "Subestacao do ativo nao encontrada")

    sigla = sigla_por_subestacao(ativo.id_subestacao)
    numero_os, numero_apr = gerar_numero_os_atendimento_ss(
        db,
        sigla,
        ativo.codigo_ativo,
    )

    nova_os = OrdemServico(
        numero_os=numero_os,
        numero_apr=numero_apr,
        numero_ss=ss.numero_ss,
        id_subestacao=ativo.id_subestacao,
        id_ativo=ativo.id_ativo,
        especie=especie_documento_por_ativo(ativo) or ativo.especie,
        instalacao=subestacao.nome,
        localizacao=ativo.bay or ss.localizacao,
        complemento=ativo.fase or ss.complemento,
        origens=f"SS {ss.numero_ss}",
        defeito=ss.descricao_problema,
        esquema_servicos=ss.esquema_servico,
        prioridade=ss.prioridade,
        responsavel=ss.equipe,
        responsavel_manutencao=ss.equipe,
        data_abertura_ss=ss.data_hora_abertura or datetime.now(),
        descricao_servicos=ss.descricao_problema,
        causa_primaria=ss.causa,
        causa_secundaria=ss.causa_secundaria,
        centro_custos=ss.centro_custo,
        status="PROGRAMADA",
    )

    ss.status = "PROGRAMADA"
    ss.numero_os = nova_os.numero_os

    db.add(nova_os)
    db.commit()
    db.refresh(nova_os)
    db.refresh(ss)

    return {
        "message": "SS atendida e OS criada com sucesso",
        "ss": ss,
        "os": nova_os,
    }


@router.get("/{id_ss}/download")
def download_ss(id_ss: int, db: Session = Depends(get_db)):
    ss = db.query(SolicitacaoServico).filter(
        SolicitacaoServico.id == id_ss
    ).first()

    if not ss:
        raise HTTPException(404, "SS nao encontrada")

    ativo = db.query(Ativo).filter(
        Ativo.id_ativo == ss.id_ativo
    ).first() if ss.id_ativo else None

    contexto = montar_contexto_ss(ss, ativo)

    pasta_saida = "saida"
    os.makedirs(pasta_saida, exist_ok=True)

    nome_arquivo = nome_arquivo_seguro(f"{ss.numero_ss}.xlsx")
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    gerar_xlsx(
        modelo="modelos/MODELO_SS.xlsx",
        destino=caminho_saida,
        contexto=contexto,
        mapeamento=MAPEAMENTO_CELULAS,
    )

    return FileResponse(
        path=caminho_saida,
        filename=nome_arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

