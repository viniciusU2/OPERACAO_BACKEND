import re
import unicodedata
from datetime import datetime
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy.orm import Session, selectinload

from models.Ativo import Ativo
from models.SS_models import SolicitacaoServico
from models.instalacao_models import Subestacao
from models.problemas_tipicos_models import ProblemaTipico
from problemas_tipicos.schemas import SSProblemaIn
from problemas_tipicos.service import sincronizar_ss


COLUNAS = {
    "id ativo": "id_ativo",
    "id do ativo": "id_ativo",
    "id_ativo": "id_ativo",
    "data solicitacao": "data_hora_solicitacao",
    "data abertura": "data_hora_abertura",
    "data limite": "data_hora_limite",
    "solicitante": "solicitante",
    "matricula": "matricula",
    "funcao": "funcao",
    "telefone": "telefone",
    "email": "email",
    "orgao": "orgao",
    "instalacao": "instalacao",
    "localizacao": "localizacao",
    "complemento": "complemento",
    "ativo": "ativo",
    "codigo ativo": "codigo_ativo",
    "tipo ativo": "tipo_ativo",
    "fase": "fase",
    "esquema servico": "esquema_servico",
    "centro custo": "centro_custo",
    "causa": "causa",
    "causa secundaria": "causa_secundaria",
    "equipe": "equipe",
    "descricao problema": "descricao_problema",
    "problema tipico": "problemas_tipicos",
    "problemas tipicos": "problemas_tipicos",
    "problemas identificados": "problemas_tipicos",
    "criticidade identificada": "criticidades_identificadas",
    "criticidades identificadas": "criticidades_identificadas",
    "observacao problema": "observacoes_problemas",
    "observacoes problemas": "observacoes_problemas",
    "problema confirmado": "problemas_confirmados",
    "prioridade": "prioridade",
    "status": "status",
}


def normalizar(valor) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto.strip().lower())


def texto(valor):
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    resultado = str(valor).strip()
    return resultado or None


def lista_texto(valor, separador=r"[;\n]+"):
    conteudo = texto(valor)
    return [item.strip() for item in re.split(separador, conteudo or "") if item.strip()]

def valor_booleano(valor) -> bool:
    return normalizar(valor) in {"1", "sim", "s", "true", "verdadeiro", "confirmado"}


def data_excel(valor, campo: str):
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, (int, float)):
        try:
            return from_excel(valor)
        except (TypeError, ValueError, OverflowError):
            pass
    valor_texto = str(valor).strip()
    for formato in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(valor_texto, formato)
        except ValueError:
            pass
    raise ValueError(f"{campo} possui data invalida: {valor_texto}")


def prioridade_ss(valor):
    prioridade = texto(valor)
    if not prioridade:
        return "NIVEL_3"
    match = re.search(r"NIVEL\s*[_-]?\s*(\d+)", prioridade, re.IGNORECASE)
    if match:
        return f"NIVEL_{match.group(1)}"
    return prioridade[:20]


def sigla_ss_subestacao(subestacao: Subestacao) -> str:
    sigla_base = subestacao.sigla or "GERAL"
    circuito = re.search(r"\bC\s*[-_]?\s*(\d+)\b", subestacao.nome or "", re.IGNORECASE)
    if circuito:
        return f"{sigla_base}-C{circuito.group(1)}"
    return sigla_base


def importar_planilha_ss(
    conteudo: bytes,
    emissor: str | None,
    db: Session,
    gerar_numero_ss,
):
    try:
        workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        worksheet = workbook["SS"] if "SS" in workbook.sheetnames else workbook.active
        linhas = worksheet.iter_rows(values_only=True)
        cabecalhos = next(linhas, None)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Nao foi possivel ler a planilha: {exc}")

    if not cabecalhos:
        raise HTTPException(status_code=400, detail="A planilha nao possui cabecalho.")

    campos = [COLUNAS.get(normalizar(cabecalho)) for cabecalho in cabecalhos]
    obrigatorios = {"solicitante", "esquema_servico"}
    ausentes = sorted(obrigatorios - {campo for campo in campos if campo})
    if ausentes:
        raise HTTPException(
            status_code=400,
            detail=f"Colunas obrigatorias ausentes: {', '.join(ausentes)}",
        )
    campos_presentes = {campo for campo in campos if campo}
    if "descricao_problema" not in campos_presentes and "problemas_tipicos" not in campos_presentes:
        raise HTTPException(status_code=400, detail="Informe a coluna Descricao Problema ou Problemas Tipicos.")

    subestacoes = db.query(Subestacao).all()
    ativos = db.query(Ativo).options(selectinload(Ativo.tipo_ativo)).all()
    subestacoes_por_chave = {}
    for subestacao in subestacoes:
        subestacoes_por_chave[normalizar(subestacao.nome)] = subestacao
        if subestacao.sigla:
            subestacoes_por_chave[normalizar(subestacao.sigla)] = subestacao

    criadas = []
    erros = []
    total_linhas = 0

    for numero_linha, valores in enumerate(linhas, start=2):
        if not any(valor not in (None, "") for valor in valores):
            continue
        total_linhas += 1
        try:
            with db.begin_nested():
                dados = {
                    campo: valores[indice] if indice < len(valores) else None
                    for indice, campo in enumerate(campos)
                    if campo
                }
                id_ativo_informado = texto(dados.get("id_ativo"))
                instalacao = texto(dados.get("instalacao"))
                codigo_ativo = texto(dados.get("codigo_ativo") or dados.get("ativo"))
                solicitante = texto(dados.get("solicitante"))
                esquema = texto(dados.get("esquema_servico"))
                titulos_problemas = lista_texto(dados.get("problemas_tipicos"))
                descricao = texto(dados.get("descricao_problema")) or "; ".join(titulos_problemas) or None

                faltantes = [
                    nome for nome, valor in (
                        ("Solicitante", solicitante),
                        ("Esquema Servico", esquema),
                        ("Descricao Problema ou Problemas Tipicos", descricao),
                    ) if not valor
                ]
                if faltantes:
                    raise ValueError(f"campos obrigatorios vazios: {', '.join(faltantes)}")

                if id_ativo_informado:
                    try:
                        id_ativo = int(float(id_ativo_informado.replace(",", ".")))
                    except ValueError:
                        raise ValueError(f"ID Ativo invalido: {id_ativo_informado}")
                    candidatos = [ativo for ativo in ativos if ativo.id_ativo == id_ativo]
                    if not candidatos:
                        raise ValueError(f"ID Ativo nao encontrado: {id_ativo}")
                    subestacao = next(
                        (sub for sub in subestacoes if sub.id_subestacao == candidatos[0].id_subestacao),
                        None,
                    )
                else:
                    if not instalacao or not codigo_ativo:
                        raise ValueError("informe ID Ativo ou Instalacao e Codigo Ativo")
                    subestacao = subestacoes_por_chave.get(normalizar(instalacao))
                    if not subestacao:
                        raise ValueError(f"instalacao nao encontrada: {instalacao}")
                    candidatos = [
                        ativo for ativo in ativos
                        if ativo.id_subestacao == subestacao.id_subestacao
                        and normalizar(ativo.codigo_ativo) == normalizar(codigo_ativo)
                    ]
                    fase = texto(dados.get("fase"))
                    if fase:
                        candidatos = [a for a in candidatos if normalizar(a.fase) == normalizar(fase)]
                    tipo_ativo = texto(dados.get("tipo_ativo"))
                    if tipo_ativo:
                        candidatos = [
                            a for a in candidatos
                            if a.tipo_ativo and normalizar(a.tipo_ativo.nome) == normalizar(tipo_ativo)
                        ]

                if not candidatos:
                    detalhe = f" / fase {fase}" if not id_ativo_informado and fase else ""
                    raise ValueError(f"ativo nao encontrado: {codigo_ativo}{detalhe}")
                if len(candidatos) > 1:
                    raise ValueError(f"ativo ambiguo: {codigo_ativo}; informe Fase e Tipo Ativo")

                ativo = candidatos[0]
                problemas_ss = []
                if titulos_problemas:
                    disponiveis = db.query(ProblemaTipico).filter(
                        ProblemaTipico.id_tipo_ativo == ativo.id_tipo_ativo,
                        ProblemaTipico.ativo.is_(True),
                    ).all()
                    por_titulo = {normalizar(item.titulo): item for item in disponiveis}
                    criticidades = lista_texto(dados.get("criticidades_identificadas"))
                    observacoes = lista_texto(dados.get("observacoes_problemas"), r"[|\n]+")
                    confirmacoes = lista_texto(dados.get("problemas_confirmados"))
                    for indice, titulo in enumerate(titulos_problemas):
                        problema = por_titulo.get(normalizar(titulo))
                        if not problema:
                            raise ValueError(f"problema tipico nao encontrado para {ativo.tipo_ativo.nome}: {titulo}")
                        problemas_ss.append(SSProblemaIn(
                            id_problema=problema.id_problema,
                            criticidade_identificada=(criticidades[indice].upper() if indice < len(criticidades) else problema.criticidade_padrao),
                            observacao=(observacoes[indice] if indice < len(observacoes) else None),
                            confirmado=(valor_booleano(confirmacoes[indice]) if indice < len(confirmacoes) else False),
                        ))

                nova_ss = SolicitacaoServico(
                    numero_ss=gerar_numero_ss(db, sigla_ss_subestacao(subestacao)),
                    data_hora_solicitacao=data_excel(dados.get("data_hora_solicitacao"), "Data Solicitacao"),
                    data_hora_abertura=data_excel(dados.get("data_hora_abertura"), "Data Abertura"),
                    data_hora_limite=data_excel(dados.get("data_hora_limite"), "Data Limite"),
                    solicitante=solicitante,
                    matricula=texto(dados.get("matricula")),
                    funcao=texto(dados.get("funcao")),
                    telefone=texto(dados.get("telefone")),
                    email=texto(dados.get("email")),
                    orgao=texto(dados.get("orgao")),
                    instalacao=subestacao.nome,
                    localizacao=texto(dados.get("localizacao")) or ativo.bay,
                    complemento=texto(dados.get("complemento")) or ativo.fase,
                    id_ativo=ativo.id_ativo,
                    id_grupo_ativo=ativo.id_grupo_ativo,
                    id_funcao_operacao=ativo.id_funcao_operacao,
                    escopo_ativo="FASE",
                    esquema_servico=esquema,
                    centro_custo=texto(dados.get("centro_custo")),
                    causa=texto(dados.get("causa")),
                    causa_secundaria=texto(dados.get("causa_secundaria")),
                    equipe=texto(dados.get("equipe")),
                    descricao_problema=descricao,
                    prioridade=prioridade_ss(dados.get("prioridade")),
                    status=texto(dados.get("status")) or "ABERTA",
                    emissor=emissor,
                )
                db.add(nova_ss)
                db.flush()
                sincronizar_ss(db, nova_ss, problemas_ss)
                criada = {
                    "linha": numero_linha,
                    "id_ss": nova_ss.id,
                    "numero_ss": nova_ss.numero_ss,
                    "codigo_ativo": ativo.codigo_ativo,
                }
            criadas.append(criada)
        except Exception as exc:
            erros.append({"linha": numero_linha, "erro": str(exc)})

    if not total_linhas:
        raise HTTPException(status_code=400, detail="A planilha nao possui linhas de dados.")

    db.commit()
    return {
        "total_linhas": total_linhas,
        "total_criadas": len(criadas),
        "total_erros": len(erros),
        "criadas": criadas,
        "erros": erros,
    }
