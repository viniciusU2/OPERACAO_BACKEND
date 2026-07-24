import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from auth.dependencies import require_roles
from auth.dependencies import garantir_colunas_usuarios
from database import get_db
from models.auth_models import Usuario
from models.instalacao_models import Subestacao
from models.sobreaviso_models import (
    SobreavisoColaborador,
    SobreavisoEquipe,
    SobreavisoHistorico,
    SobreavisoIntervalo,
    SobreavisoPeriodo,
    SobreavisoSolicitacaoAjuste,
)
from Sobreaviso.schemas import (
    ColaboradorSobreavisoCreate,
    ColaboradorSobreavisoResponse,
    ColaboradorSobreavisoUpdate,
    EquipeSobreavisoCreate,
    EquipeSobreavisoResponse,
    EquipeSobreavisoUpdate,
    HistoricoSobreavisoResponse,
    ResumoSobreavisoResponse,
    SobreavisoCreate,
    SobreavisoResponse,
    SobreavisoIntervaloCreate,
    SobreavisoIntervaloResponse,
    SobreavisoUpdate,
    SolicitacaoAjusteCreate,
    SolicitacaoAjusteResponse,
)

router = APIRouter(prefix="/sobreaviso", tags=["Sobreaviso"])

STATUS_VALIDOS = {"PLANEJADO", "PENDENTE", "APROVADO", "REPROVADO", "CANCELADO"}
ORIGENS_VALIDAS = {"ADMIN", "GESTOR", "COLABORADOR", "IMPORTACAO"}
STATUS_INATIVOS_PARA_CONFLITO = {"CANCELADO", "REPROVADO"}
DIAS_SEMANA = ["SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"]
MESES_PT = [
    "JANEIRO", "FEVEREIRO", "MARCO", "ABRIL", "MAIO", "JUNHO",
    "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
]


def calcular_total_horas(inicio: datetime, fim: datetime) -> Decimal:
    if fim <= inicio:
        raise HTTPException(status_code=400, detail="O fim deve ser maior que o inicio")
    segundos = (fim - inicio).total_seconds()
    return Decimal(str(round(segundos / 3600, 2)))


def validar_intervalos(
    inicio_programado: datetime,
    fim_programado: datetime,
    intervalos: list[SobreavisoIntervaloCreate],
):
    """Valida a linha do tempo e devolve os totais por tipo."""
    ordenados = sorted(intervalos, key=lambda item: item.inicio)
    anterior = None
    total_sobreaviso = Decimal("0")
    total_atendimento = Decimal("0")
    for item in ordenados:
        if item.inicio < inicio_programado or item.fim > fim_programado:
            raise HTTPException(status_code=400, detail="Intervalo fora do periodo programado")
        if anterior and item.inicio < anterior.fim:
            raise HTTPException(status_code=400, detail="Nao sao permitidos intervalos sobrepostos")
        horas = calcular_total_horas(item.inicio, item.fim)
        if item.tipo == "SOBREAVISO":
            total_sobreaviso += horas
        else:
            total_atendimento += horas
        anterior = item
    return ordenados, total_sobreaviso, total_atendimento


def substituir_intervalos(
    db: Session,
    sobreaviso: SobreavisoPeriodo,
    intervalos: list[SobreavisoIntervaloCreate],
    usuario: Usuario,
):
    ordenados, total_sobreaviso, total_atendimento = validar_intervalos(
        sobreaviso.inicio, sobreaviso.fim, intervalos
    )
    sobreaviso.intervalos.clear()
    for item in ordenados:
        sobreaviso.intervalos.append(SobreavisoIntervalo(
            tipo=item.tipo,
            inicio=item.inicio,
            fim=item.fim,
            id_ocorrencia=item.id_ocorrencia,
            observacao=item.observacao,
            criado_por=usuario.id,
            atualizado_por=usuario.id,
        ))
    sobreaviso.total_horas = total_sobreaviso
    sobreaviso.total_horas_atendimento = total_atendimento


def normalizar_status(status: str) -> str:
    valor = (status or "").strip().upper()
    if valor not in STATUS_VALIDOS:
        raise HTTPException(status_code=400, detail="Status de sobreaviso invalido")
    return valor


def normalizar_origem(origem: str) -> str:
    valor = (origem or "").strip().upper()
    if valor not in ORIGENS_VALIDAS:
        raise HTTPException(status_code=400, detail="Origem de sobreaviso invalida")
    return valor


def registrar_historico(
    db: Session,
    entidade: str,
    entidade_id: int,
    acao: str,
    usuario: Optional[Usuario],
    dados_anteriores: Optional[dict] = None,
    dados_novos: Optional[dict] = None,
    justificativa: Optional[str] = None,
):
    db.add(
        SobreavisoHistorico(
            entidade=entidade,
            entidade_id=entidade_id,
            acao=acao,
            dados_anteriores=json.dumps(dados_anteriores, default=str) if dados_anteriores else None,
            dados_novos=json.dumps(dados_novos, default=str) if dados_novos else None,
            justificativa=justificativa,
            alterado_por=usuario.id if usuario else None,
        )
    )


def buscar_colaborador_ou_404(
    db: Session,
    id_colaborador: int,
    exigir_ativo: bool = False,
) -> SobreavisoColaborador:
    colaborador = (
        db.query(SobreavisoColaborador)
        .filter(SobreavisoColaborador.id_colaborador == id_colaborador)
        .first()
    )
    if not colaborador:
        raise HTTPException(status_code=404, detail="Colaborador de sobreaviso nao encontrado")
    if exigir_ativo and not bool(colaborador.ativo):
        raise HTTPException(status_code=400, detail="O colaborador selecionado esta inativo")
    return colaborador


def buscar_sobreaviso_ou_404(db: Session, id_sobreaviso: int) -> SobreavisoPeriodo:
    sobreaviso = (
        db.query(SobreavisoPeriodo)
        .options(
            selectinload(SobreavisoPeriodo.colaborador),
            selectinload(SobreavisoPeriodo.intervalos),
        )
        .filter(SobreavisoPeriodo.id_sobreaviso == id_sobreaviso)
        .first()
    )
    if not sobreaviso:
        raise HTTPException(status_code=404, detail="Sobreaviso nao encontrado")
    return sobreaviso


def existe_sobreposicao(
    db: Session,
    id_colaborador: int,
    inicio: datetime,
    fim: datetime,
    ignorar_id: Optional[int] = None,
) -> bool:
    query = db.query(SobreavisoPeriodo).filter(
        SobreavisoPeriodo.id_colaborador == id_colaborador,
        SobreavisoPeriodo.inicio < fim,
        SobreavisoPeriodo.fim > inicio,
        SobreavisoPeriodo.status.notin_(STATUS_INATIVOS_PARA_CONFLITO),
    )

    if ignorar_id:
        query = query.filter(SobreavisoPeriodo.id_sobreaviso != ignorar_id)

    return db.query(query.exists()).scalar()


def aplicar_filtros_sobreaviso(
    query,
    data_inicio: Optional[datetime],
    data_fim: Optional[datetime],
    id_equipe: Optional[int],
    id_colaborador: Optional[int],
    status: Optional[str],
    busca: Optional[str],
):
    if data_inicio:
        query = query.filter(SobreavisoPeriodo.fim >= data_inicio)
    if data_fim:
        query = query.filter(SobreavisoPeriodo.inicio <= data_fim)
    if id_colaborador:
        query = query.filter(SobreavisoPeriodo.id_colaborador == id_colaborador)
    if status and status != "all":
        query = query.filter(SobreavisoPeriodo.status == normalizar_status(status))
    if id_equipe or busca:
        query = query.join(SobreavisoPeriodo.colaborador)
    if id_equipe:
        query = query.filter(SobreavisoColaborador.id_equipe == id_equipe)
    if busca:
        termo = f"%{busca}%"
        query = query.filter(
            or_(
                SobreavisoColaborador.nome.like(termo),
                SobreavisoColaborador.matricula.like(termo),
                SobreavisoPeriodo.justificativa.like(termo),
            )
        )
    return query


def formatar_hora_excel(valor: datetime) -> str:
    return valor.strftime("%H:%M")


def horas_para_texto(horas: Decimal) -> str:
    minutos = int(round(float(horas) * 60))
    return f"{minutos // 60}:{minutos % 60:02d}"


def limpar_nome_arquivo(valor: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", valor.strip())[:80] or "relatorio"


def dividir_periodo_por_dia(inicio: datetime, fim: datetime):
    atual = inicio
    while atual.date() <= fim.date():
        dia_inicio = datetime.combine(atual.date(), datetime.min.time())
        dia_fim = datetime.combine(atual.date(), datetime.max.time()).replace(microsecond=0)
        trecho_inicio = max(inicio, dia_inicio)
        trecho_fim = min(fim, dia_fim)
        if trecho_fim > trecho_inicio:
            yield trecho_inicio.date(), trecho_inicio, trecho_fim
        atual = datetime.combine(atual.date(), datetime.min.time()).replace(day=atual.day) 
        atual = datetime.fromordinal(atual.date().toordinal() + 1)


def montar_relatorio_folha_ponto(
    colaborador: SobreavisoColaborador,
    sobreavisos: list[SobreavisoPeriodo],
    data_inicio: datetime,
    data_fim: datetime,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "FOLHA DE PONTO GERAL"

    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="E2F0D9")
    title_fill = PatternFill("solid", fgColor="1F4E78")

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.merge_cells("A1:N1")
    ws["A1"] = f"FOLHA DE PONTO GERAL - {data_inicio.strftime('%d/%m/%Y')} A {data_fim.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "NOME:"
    ws["B2"] = colaborador.nome
    ws["C2"] = "MATRICULA:"
    ws["D2"] = colaborador.matricula
    ws["A3"] = "FUNCAO:"
    ws["B3"] = colaborador.cargo or ""
    ws["C3"] = "JORNADA DE TRABALHO:"
    ws["D3"] = "07:00 AS 17:00 HS"
    ws["K3"] = "MES:"
    ws["L3"] = data_fim.strftime("%m")
    ws["M3"] = "ANO:"
    ws["N3"] = data_fim.year
    ws["A4"] = "EMPRESA"
    ws["B4"] = "RIALMA TRANSMISSORA"
    ws["C4"] = "LOCALIDADE:"
    ws["D4"] = colaborador.equipe.nome if colaborador.equipe else ""
    ws["F4"] = "ESTADO:"
    ws["G4"] = "BAHIA"
    ws["A5"] = "CNPJ:"
    ws["B5"] = "03.286.850/0001-96"
    ws["C5"] = "ATIVIDADE:"
    ws["D5"] = "TRANSMISSAO DE ENERGIA ELETRICA"

    ws.merge_cells("A7:D7")
    ws["A7"] = "DIAS"
    ws.merge_cells("E7:H7")
    ws["E7"] = "EXPEDIENTE NORMAL"
    ws.merge_cells("I7:N7")
    ws["I7"] = "SOBREAVISO"
    ws["A8"] = "DIA"
    ws["B8"] = "CODIGO"
    ws["C8"] = "N."
    ws["D8"] = "DATA"
    ws["E8"] = "INICIO"
    ws["F8"] = "INTERVALO"
    ws["G8"] = "TERMINO"
    ws["H8"] = "TOTAL DIARIO"
    ws["I8"] = "INICIO"
    ws["J8"] = "INTERVALO"
    ws["K8"] = "TERMINO"
    ws["L8"] = "TOTAL DIARIO"
    ws["M8"] = "OBSERVACOES"
    ws["N8"] = "TOTAL"

    for row in range(7, 9):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill if row == 7 else header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    horas_por_dia: dict[str, list[tuple[datetime, datetime, Decimal, str]]] = {}
    for item in sobreavisos:
        periodo_inicio = max(item.inicio, data_inicio)
        periodo_fim = min(item.fim, data_fim)
        for dia, trecho_inicio, trecho_fim in dividir_periodo_por_dia(periodo_inicio, periodo_fim):
            horas = calcular_total_horas(trecho_inicio, trecho_fim)
            chave = dia.isoformat()
            horas_por_dia.setdefault(chave, []).append((trecho_inicio, trecho_fim, horas, item.justificativa or ""))

    row = 9
    dia_atual = data_inicio.date()
    while dia_atual <= data_fim.date():
        eventos = horas_por_dia.get(dia_atual.isoformat(), [])
        total_dia = sum((evento[2] for evento in eventos), Decimal("0"))

        ws.cell(row=row, column=1, value=dia_atual.day)
        ws.cell(row=row, column=2, value=DIAS_SEMANA[dia_atual.weekday()])
        ws.cell(row=row, column=3, value=dia_atual.weekday() + 1)
        ws.cell(row=row, column=4, value=dia_atual)
        ws.cell(row=row, column=8, value="0:00")

        if eventos:
            # O mesmo dia pode ter varios trechos de sobreaviso. A folha tem
            # apenas uma linha diaria, portanto exibe a faixa completa e as
            # interrupcoes na coluna INTERVALO, sem ocultar os trechos seguintes.
            inicio = eventos[0][0]
            fim = eventos[-1][1]
            lacunas = []
            for evento_anterior, proximo_evento in zip(eventos, eventos[1:]):
                if proximo_evento[0] > evento_anterior[1]:
                    lacunas.append(
                        f"{formatar_hora_excel(evento_anterior[1])}-{formatar_hora_excel(proximo_evento[0])}"
                    )
            observacoes = [evento[3] for evento in eventos if evento[3]]
            ws.cell(row=row, column=9, value=formatar_hora_excel(inicio))
            ws.cell(row=row, column=10, value=" / ".join(lacunas))
            ws.cell(row=row, column=11, value="24:00" if fim.hour == 23 and fim.minute == 59 else formatar_hora_excel(fim))
            ws.cell(row=row, column=12, value=horas_para_texto(total_dia))
            ws.cell(row=row, column=13, value=" | ".join(dict.fromkeys(observacoes)))
            ws.cell(row=row, column=14, value=horas_para_texto(total_dia))
        else:
            ws.cell(row=row, column=12, value="0:00")
            ws.cell(row=row, column=14, value="0:00")

        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=row, column=13).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=4).number_format = "dd/mm/yyyy"

        row += 1
        dia_atual = datetime.fromordinal(dia_atual.toordinal() + 1).date()

    total_geral = sum(
        (
            calcular_total_horas(max(item.inicio, data_inicio), min(item.fim, data_fim))
            for item in sobreavisos
            if item.status != "CANCELADO" and min(item.fim, data_fim) > max(item.inicio, data_inicio)
        ),
        Decimal("0"),
    )
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=14, value=horas_para_texto(total_geral))
    for col in range(1, 15):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="Reconheco a exatidao destas Anotacoes:")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="____________________________________________________________ Assinatura do Funcionario")
    ws.cell(row=row, column=8, value="____________________________________________________________ Assinatura do Responsavel")

    ws.freeze_panes = "A9"
    return wb


def montar_relatorio_escala_geral(
    subestacoes: list[Subestacao],
    colaboradores: list[SobreavisoColaborador],
    sobreavisos: list[SobreavisoPeriodo],
    data_inicio: datetime,
    data_fim: datetime,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "ESCALA GERAL"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E5"

    dias = []
    dia = data_inicio.date()
    ultimo_dia = data_fim.date()
    while dia <= ultimo_dia:
        dias.append(dia)
        dia += timedelta(days=1)

    primeira_coluna_dia = 5
    coluna_final = primeira_coluna_dia + len(dias) - 1
    ultima_coluna = get_column_letter(coluna_final)

    azul = "17365D"
    azul_claro = "DCE6F1"
    cinza = "D9E1F2"
    cinza_secao = "E7E6E6"
    verde_fim_semana = "C6EFCE"
    azul_escala = "DDEBF7"
    vermelho = "C00000"
    branco = "FFFFFF"
    borda_fina = Side(style="thin", color="707070")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 15
    for indice in range(primeira_coluna_dia, coluna_final + 1):
        ws.column_dimensions[get_column_letter(indice)].width = 11

    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=coluna_final)
    ws.cell(1, 4, "ESCALA GERAL DE SOBREAVISO - RIALMA TRANSMISSORA DE ENERGIA V")
    ws.cell(1, 4).font = Font(name="Arial", size=16, bold=True, color=branco)
    ws.cell(1, 4).fill = PatternFill("solid", fgColor=azul)
    ws.cell(1, 4).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 27

    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=coluna_final)
    ws.cell(2, 4, f"COMPETENCIA: {data_inicio.strftime('%d/%m/%Y')} A {data_fim.strftime('%d/%m/%Y')}")
    ws.cell(2, 4).font = Font(name="Arial", size=10, bold=True, color=azul)
    ws.cell(2, 4).fill = PatternFill("solid", fgColor=azul_claro)
    ws.cell(2, 4).alignment = Alignment(horizontal="center")

    ws.merge_cells("A1:C2")
    ws["A1"] = "RIALMA S.A."
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    periodos_por_colaborador: dict[int, list[SobreavisoPeriodo]] = {}
    for item in sobreavisos:
        periodos_por_colaborador.setdefault(item.id_colaborador, []).append(item)

    colaboradores_por_subestacao: dict[Optional[int], list[SobreavisoColaborador]] = {}
    for colaborador in colaboradores:
        colaboradores_por_subestacao.setdefault(colaborador.id_subestacao, []).append(colaborador)

    grupos = [(sub.id_subestacao, sub.nome, sub.sigla or "SE") for sub in subestacoes]
    if colaboradores_por_subestacao.get(None):
        grupos.append((None, "SEM SUBESTACAO DEFINIDA", "-"))

    linha = 4
    for id_subestacao, nome_subestacao, sigla_subestacao in grupos:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=coluna_final)
        ws.cell(linha, 1, f"{MESES_PT[data_fim.month - 1]} - {nome_subestacao.upper()}")
        ws.cell(linha, 1).font = Font(name="Arial", size=10, bold=True, color=azul)
        ws.cell(linha, 1).fill = PatternFill("solid", fgColor=cinza_secao)
        ws.cell(linha, 1).alignment = Alignment(horizontal="left")
        ws.cell(linha, 1).border = borda
        linha += 1

        cabecalhos = ["Colaborador", "SE", "Telefone", "Equipe"]
        for coluna, texto in enumerate(cabecalhos, start=1):
            celula = ws.cell(linha, coluna, texto)
            celula.fill = PatternFill("solid", fgColor=cinza)
            celula.font = Font(name="Arial", size=8, bold=True, color=azul)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celula.border = borda

        for deslocamento, data_dia in enumerate(dias):
            coluna = primeira_coluna_dia + deslocamento
            celula = ws.cell(linha, coluna, f"{data_dia.strftime('%d/%m')}\n{DIAS_SEMANA[data_dia.weekday()]}")
            celula.fill = PatternFill(
                "solid",
                fgColor=verde_fim_semana if data_dia.weekday() >= 5 else cinza,
            )
            celula.font = Font(name="Arial", size=7, bold=True, color=azul)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celula.border = borda

        ws.row_dimensions[linha].height = 28
        linha += 1

        colaboradores_grupo = colaboradores_por_subestacao.get(id_subestacao, [])
        if not colaboradores_grupo:
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
            ws.cell(linha, 1, "Nenhum colaborador ativo vinculado a esta subestacao")
            ws.cell(linha, 1).font = Font(name="Arial", size=8, italic=True, color="666666")
            ws.cell(linha, 1).alignment = Alignment(horizontal="left")
            for coluna in range(1, coluna_final + 1):
                ws.cell(linha, coluna).border = borda
            linha += 1
        else:
            for colaborador in colaboradores_grupo:
                ws.cell(linha, 1, colaborador.nome)
                ws.cell(linha, 2, sigla_subestacao)
                ws.cell(linha, 3, colaborador.telefone or "-")
                ws.cell(linha, 4, colaborador.equipe.nome if colaborador.equipe else "-")
                for deslocamento, data_dia in enumerate(dias):
                    coluna = primeira_coluna_dia + deslocamento
                    inicio_dia = datetime.combine(data_dia, datetime.min.time())
                    fim_dia = inicio_dia + timedelta(days=1)
                    intervalos = []

                    for periodo in periodos_por_colaborador.get(colaborador.id_colaborador, []):
                        inicio = max(periodo.inicio, inicio_dia, data_inicio)
                        fim = min(periodo.fim, fim_dia, data_fim)
                        if fim <= inicio:
                            continue
                        fim_texto = "24:00" if fim == fim_dia else fim.strftime("%H:%M")
                        intervalos.append(f"{inicio.strftime('%H:%M')}-{fim_texto}")

                    celula = ws.cell(linha, coluna, "\n".join(intervalos))
                    celula.fill = PatternFill(
                        "solid",
                        fgColor=verde_fim_semana if data_dia.weekday() >= 5 else azul_escala if intervalos else branco,
                    )
                    celula.font = Font(name="Arial", size=7, color=azul, bold=bool(intervalos))
                    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    celula.border = borda

                for coluna in range(1, coluna_final + 1):
                    celula = ws.cell(linha, coluna)
                    celula.border = borda
                    if coluna <= 4:
                        celula.font = Font(name="Arial", size=8)
                        celula.alignment = Alignment(horizontal="left" if coluna in (1, 3, 4) else "center", vertical="center")
                ws.row_dimensions[linha].height = 26
                linha += 1

        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=coluna_final)
        ws.cell(linha, 1, "OBSERVACOES:")
        ws.cell(linha, 1).font = Font(name="Arial", size=8, bold=True, color=vermelho)
        ws.cell(linha, 1).border = borda
        linha += 1
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=coluna_final)
        ws.cell(
            linha,
            1,
            "1 - Horarios apresentados conforme os registros aprovados, pendentes ou planejados do sistema. "
            "2 - Registros cancelados e reprovados nao integram esta escala.",
        )
        ws.cell(linha, 1).font = Font(name="Arial", size=7, color="404040")
        ws.cell(linha, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(linha, 1).border = borda
        ws.row_dimensions[linha].height = 27
        linha += 2

    ws.auto_filter.ref = f"A5:{ultima_coluna}5"
    ws.print_area = f"A1:{ultima_coluna}{max(linha - 1, 5)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.sheet_view.zoomScale = 70
    return wb


def garantir_equipes_iniciais(db: Session):
    if db.query(SobreavisoEquipe).count() > 0:
        return

    for nome, descricao in (
        ("Operacao", "Equipe de operacao em campo"),
        ("Manutencao", "Equipe de manutencao eletrica"),
        ("Suporte", "Equipe de apoio tecnico"),
    ):
        db.add(SobreavisoEquipe(nome=nome, descricao=descricao, ativo=1))
    db.commit()


def garantir_colunas_sobreaviso(db: Session):
    existe = db.execute(
        text("SHOW COLUMNS FROM sobreaviso_colaborador LIKE :coluna"),
        {"coluna": "id_subestacao"},
    ).first()
    if not existe:
        db.execute(text("ALTER TABLE sobreaviso_colaborador ADD COLUMN id_subestacao INT NULL"))
    total_atendimento = db.execute(
        text("SHOW COLUMNS FROM sobreaviso_periodo LIKE :coluna"),
        {"coluna": "total_horas_atendimento"},
    ).first()
    if not total_atendimento:
        db.execute(text("ALTER TABLE sobreaviso_periodo ADD COLUMN total_horas_atendimento DECIMAL(10,2) NOT NULL DEFAULT 0"))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS sobreaviso_intervalo (
            id_intervalo INT AUTO_INCREMENT PRIMARY KEY,
            id_sobreaviso INT NOT NULL,
            tipo VARCHAR(20) NOT NULL,
            inicio DATETIME NOT NULL,
            fim DATETIME NOT NULL,
            id_ocorrencia INT NULL,
            observacao TEXT NULL,
            criado_por INT NULL,
            atualizado_por INT NULL,
            criado_em DATETIME NULL,
            atualizado_em DATETIME NULL,
            INDEX ix_sobreaviso_intervalo_sobreaviso (id_sobreaviso),
            INDEX ix_sobreaviso_intervalo_inicio (inicio),
            CONSTRAINT fk_sobreaviso_intervalo_periodo FOREIGN KEY (id_sobreaviso)
                REFERENCES sobreaviso_periodo(id_sobreaviso) ON DELETE CASCADE
        ) ENGINE=InnoDB
    """))
    db.execute(text("""
        INSERT INTO sobreaviso_intervalo (id_sobreaviso, tipo, inicio, fim, criado_por, atualizado_por, criado_em)
        SELECT p.id_sobreaviso, 'SOBREAVISO', p.inicio, p.fim, p.criado_por, p.atualizado_por, p.criado_em
        FROM sobreaviso_periodo p
        LEFT JOIN sobreaviso_intervalo i ON i.id_sobreaviso = p.id_sobreaviso
        WHERE i.id_intervalo IS NULL
    """))
    db.commit()


def equipe_padrao_id(db: Session) -> int:
    garantir_equipes_iniciais(db)
    equipe = db.query(SobreavisoEquipe).order_by(SobreavisoEquipe.id_equipe).first()
    if not equipe:
        raise HTTPException(status_code=500, detail="Nao foi possivel criar equipe padrao")
    return equipe.id_equipe


def sincronizar_colaboradores_usuarios(db: Session):
    garantir_colunas_usuarios(db)
    garantir_colunas_sobreaviso(db)
    id_equipe = equipe_padrao_id(db)
    usuarios_ativos = db.query(Usuario).filter(Usuario.ativo == True).all()
    resumo = {
        "usuarios_ativos": len(usuarios_ativos),
        "ignorados_admin": 0,
        "criados": 0,
        "atualizados": 0,
    }

    for usuario in usuarios_ativos:
        role = (usuario.role or "usuario").strip().lower()
        if role == "admin":
            resumo["ignorados_admin"] += 1
            continue

        colaborador = (
            db.query(SobreavisoColaborador)
            .filter(SobreavisoColaborador.id_usuario == usuario.id)
            .first()
        )
        matricula_usuario = f"USR-{usuario.id}"

        if colaborador:
            colaborador.nome = usuario.nome
            colaborador.email = usuario.email
            colaborador.matricula = colaborador.matricula or matricula_usuario
            colaborador.cargo = colaborador.cargo or usuario.role
            colaborador.ativo = 1
            if usuario.id_subestacao_padrao and colaborador.id_subestacao != usuario.id_subestacao_padrao:
                colaborador.id_subestacao = usuario.id_subestacao_padrao
            resumo["atualizados"] += 1
            continue

        colaborador = (
            db.query(SobreavisoColaborador)
            .filter(SobreavisoColaborador.matricula == matricula_usuario)
            .first()
        )

        if not colaborador:
            colaborador = (
                db.query(SobreavisoColaborador)
                .filter(func.lower(SobreavisoColaborador.email) == usuario.email.lower())
                .first()
            )

        if colaborador:
            colaborador.id_usuario = usuario.id
            colaborador.matricula = colaborador.matricula or matricula_usuario
            colaborador.id_equipe = colaborador.id_equipe or id_equipe
            colaborador.id_subestacao = colaborador.id_subestacao or usuario.id_subestacao_padrao
            colaborador.nome = usuario.nome
            colaborador.email = usuario.email
            colaborador.cargo = colaborador.cargo or usuario.role
            colaborador.ativo = 1
            resumo["atualizados"] += 1
            continue

        db.add(
            SobreavisoColaborador(
                id_usuario=usuario.id,
                id_equipe=id_equipe,
                id_subestacao=usuario.id_subestacao_padrao,
                nome=usuario.nome,
                matricula=matricula_usuario,
                email=usuario.email,
                cargo=usuario.role,
                ativo=1,
            )
        )
        resumo["criados"] += 1

    db.commit()
    return resumo


@router.get("/equipes", response_model=list[EquipeSobreavisoResponse])
def listar_equipes(
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    garantir_equipes_iniciais(db)
    return db.query(SobreavisoEquipe).order_by(SobreavisoEquipe.nome).all()


@router.post("/equipes", response_model=EquipeSobreavisoResponse, status_code=201)
def criar_equipe(
    dados: EquipeSobreavisoCreate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor")),
):
    equipe = SobreavisoEquipe(
        nome=dados.nome,
        descricao=dados.descricao,
        ativo=1 if dados.ativo else 0,
    )
    db.add(equipe)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Ja existe equipe com este nome")

    registrar_historico(db, "EQUIPE", equipe.id_equipe, "CRIACAO", usuario, dados_novos=dados.model_dump())
    db.commit()
    db.refresh(equipe)
    return equipe


@router.put("/equipes/{id_equipe}", response_model=EquipeSobreavisoResponse)
def atualizar_equipe(
    id_equipe: int,
    dados: EquipeSobreavisoUpdate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor")),
):
    equipe = db.query(SobreavisoEquipe).filter(SobreavisoEquipe.id_equipe == id_equipe).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Equipe nao encontrada")

    anteriores = {"nome": equipe.nome, "descricao": equipe.descricao, "ativo": bool(equipe.ativo)}
    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(equipe, campo, 1 if campo == "ativo" and valor else 0 if campo == "ativo" else valor)

    registrar_historico(
        db,
        "EQUIPE",
        id_equipe,
        "EDICAO",
        usuario,
        dados_anteriores=anteriores,
        dados_novos=dados.model_dump(exclude_unset=True),
    )
    db.commit()
    db.refresh(equipe)
    return equipe


@router.get("/colaboradores", response_model=list[ColaboradorSobreavisoResponse])
def listar_colaboradores(
    id_equipe: Optional[int] = None,
    ativo: Optional[bool] = None,
    busca: Optional[str] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    sincronizar_colaboradores_usuarios(db)
    query = (
        db.query(SobreavisoColaborador)
        .outerjoin(Usuario, SobreavisoColaborador.id_usuario == Usuario.id)
        .filter(
            or_(
                SobreavisoColaborador.id_usuario.is_(None),
                func.lower(func.coalesce(Usuario.role, "usuario")) != "admin",
            )
        )
    )
    if id_equipe:
        query = query.filter(SobreavisoColaborador.id_equipe == id_equipe)
    if ativo is not None:
        query = query.filter(SobreavisoColaborador.ativo == (1 if ativo else 0))
    if busca:
        termo = f"%{busca}%"
        query = query.filter(
            or_(
                SobreavisoColaborador.nome.like(termo),
                SobreavisoColaborador.matricula.like(termo),
                SobreavisoColaborador.email.like(termo),
            )
        )
    return query.order_by(SobreavisoColaborador.nome).all()


@router.post("/sincronizar-colaboradores")
@router.post("/colaboradores/sincronizar")
def sincronizar_colaboradores(
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    resumo = sincronizar_colaboradores_usuarios(db)
    total_colaboradores = db.query(SobreavisoColaborador).count()
    return {
        "message": "Sincronizacao de colaboradores concluida",
        "total_colaboradores": total_colaboradores,
        **resumo,
    }


@router.post("/colaboradores", response_model=ColaboradorSobreavisoResponse, status_code=201)
def criar_colaborador(
    dados: ColaboradorSobreavisoCreate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor")),
):
    garantir_colunas_sobreaviso(db)
    buscar_equipe = db.query(SobreavisoEquipe).filter(SobreavisoEquipe.id_equipe == dados.id_equipe).first()
    if not buscar_equipe:
        raise HTTPException(status_code=404, detail="Equipe nao encontrada")
    if dados.id_subestacao:
        subestacao = db.query(Subestacao).filter(Subestacao.id_subestacao == dados.id_subestacao).first()
        if not subestacao:
            raise HTTPException(status_code=404, detail="Subestacao nao encontrada")

    colaborador = SobreavisoColaborador(
        **dados.model_dump(exclude={"ativo"}),
        ativo=1 if dados.ativo else 0,
    )
    db.add(colaborador)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Ja existe colaborador com esta matricula")

    registrar_historico(
        db,
        "COLABORADOR",
        colaborador.id_colaborador,
        "CRIACAO",
        usuario,
        dados_novos=dados.model_dump(),
    )
    db.commit()
    if dados.id_usuario and dados.id_subestacao:
        usuario_vinculado = db.query(Usuario).filter(Usuario.id == dados.id_usuario).first()
        if usuario_vinculado:
            usuario_vinculado.id_subestacao_padrao = dados.id_subestacao
            db.commit()
    db.refresh(colaborador)
    return colaborador


@router.put("/colaboradores/{id_colaborador}", response_model=ColaboradorSobreavisoResponse)
def atualizar_colaborador(
    id_colaborador: int,
    dados: ColaboradorSobreavisoUpdate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor")),
):
    garantir_colunas_sobreaviso(db)
    colaborador = buscar_colaborador_ou_404(db, id_colaborador)
    if dados.id_subestacao:
        subestacao = db.query(Subestacao).filter(Subestacao.id_subestacao == dados.id_subestacao).first()
        if not subestacao:
            raise HTTPException(status_code=404, detail="Subestacao nao encontrada")
    anteriores = {
        "nome": colaborador.nome,
        "matricula": colaborador.matricula,
        "email": colaborador.email,
        "id_equipe": colaborador.id_equipe,
        "id_subestacao": colaborador.id_subestacao,
        "ativo": bool(colaborador.ativo),
    }

    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(
            colaborador,
            campo,
            1 if campo == "ativo" and valor else 0 if campo == "ativo" else valor,
        )
    colaborador.atualizado_em = datetime.utcnow()

    registrar_historico(
        db,
        "COLABORADOR",
        id_colaborador,
        "EDICAO",
        usuario,
        dados_anteriores=anteriores,
        dados_novos=dados.model_dump(exclude_unset=True),
    )
    if colaborador.id_usuario and colaborador.id_subestacao:
        usuario_vinculado = db.query(Usuario).filter(Usuario.id == colaborador.id_usuario).first()
        if usuario_vinculado:
            usuario_vinculado.id_subestacao_padrao = colaborador.id_subestacao
    db.commit()
    db.refresh(colaborador)
    return colaborador


@router.get("/", response_model=list[SobreavisoResponse])
def listar_sobreavisos(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    id_equipe: Optional[int] = None,
    id_colaborador: Optional[int] = None,
    status: Optional[str] = None,
    busca: Optional[str] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    sincronizar_colaboradores_usuarios(db)
    query = db.query(SobreavisoPeriodo).options(selectinload(SobreavisoPeriodo.colaborador))
    query = aplicar_filtros_sobreaviso(query, data_inicio, data_fim, id_equipe, id_colaborador, status, busca)
    return query.order_by(SobreavisoPeriodo.inicio.desc(), SobreavisoPeriodo.id_sobreaviso.desc()).all()


@router.get("/calendario", response_model=list[SobreavisoResponse])
def listar_calendario(
    data_inicio: datetime,
    data_fim: datetime,
    id_equipe: Optional[int] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    query = db.query(SobreavisoPeriodo).options(selectinload(SobreavisoPeriodo.colaborador))
    query = aplicar_filtros_sobreaviso(query, data_inicio, data_fim, id_equipe, None, None, None)
    return query.order_by(SobreavisoPeriodo.inicio).all()


@router.get("/pendentes", response_model=list[SobreavisoResponse])
def listar_pendentes(
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    return (
        db.query(SobreavisoPeriodo)
        .options(selectinload(SobreavisoPeriodo.colaborador))
        .filter(SobreavisoPeriodo.status == "PENDENTE")
        .order_by(SobreavisoPeriodo.inicio)
        .all()
    )


@router.get("/relatorios/resumo", response_model=ResumoSobreavisoResponse)
def resumo_sobreavisos(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    id_equipe: Optional[int] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    query = db.query(SobreavisoPeriodo).options(selectinload(SobreavisoPeriodo.colaborador))
    query = aplicar_filtros_sobreaviso(query, data_inicio, data_fim, id_equipe, None, None, None)
    itens = query.all()
    total_horas = sum((item.total_horas or Decimal("0")) for item in itens if item.status != "CANCELADO")
    total_aprovadas = sum((item.total_horas or Decimal("0")) for item in itens if item.status == "APROVADO")
    return {
        "total_horas": total_horas,
        "total_aprovadas": total_aprovadas,
        "pendentes": len([item for item in itens if item.status == "PENDENTE"]),
        "planejados": len([item for item in itens if item.status == "PLANEJADO"]),
        "colaboradores": len({item.id_colaborador for item in itens}),
    }


@router.get("/historico", response_model=list[HistoricoSobreavisoResponse])
def listar_historico(
    entidade: Optional[str] = None,
    entidade_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    query = db.query(SobreavisoHistorico)
    if entidade:
        query = query.filter(SobreavisoHistorico.entidade == entidade)
    if entidade_id:
        query = query.filter(SobreavisoHistorico.entidade_id == entidade_id)
    return query.order_by(SobreavisoHistorico.criado_em.desc()).all()


@router.get("/solicitacoes-ajuste", response_model=list[SolicitacaoAjusteResponse])
def listar_solicitacoes_ajuste(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    query = db.query(SobreavisoSolicitacaoAjuste)
    if status and status != "all":
        query = query.filter(SobreavisoSolicitacaoAjuste.status == status.upper())
    return query.order_by(SobreavisoSolicitacaoAjuste.criado_em.desc()).all()


@router.get("/relatorios/folha-ponto/exportar")
def exportar_folha_ponto_sobreaviso(
    id_colaborador: int,
    data_inicio: datetime,
    data_fim: datetime,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    if data_fim < data_inicio:
        raise HTTPException(status_code=400, detail="Data final deve ser maior ou igual a data inicial")

    colaborador = buscar_colaborador_ou_404(db, id_colaborador)
    sobreavisos = (
        db.query(SobreavisoPeriodo)
        .options(selectinload(SobreavisoPeriodo.colaborador).selectinload(SobreavisoColaborador.equipe))
        .filter(
            SobreavisoPeriodo.id_colaborador == id_colaborador,
            SobreavisoPeriodo.status != "REPROVADO",
            SobreavisoPeriodo.status != "CANCELADO",
            SobreavisoPeriodo.inicio <= data_fim,
            SobreavisoPeriodo.fim >= data_inicio,
        )
        .order_by(SobreavisoPeriodo.inicio)
        .all()
    )

    wb = montar_relatorio_folha_ponto(colaborador, sobreavisos, data_inicio, data_fim)
    output_dir = Path("exports") / "sobreaviso"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = (
        f"folha_ponto_sobreaviso_{limpar_nome_arquivo(colaborador.nome)}_"
        f"{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
    )
    output_path = output_dir / filename
    wb.save(output_path)

    return FileResponse(
        path=str(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.get("/relatorios/escala-geral/exportar")
def exportar_escala_geral_sobreaviso(
    data_inicio: datetime,
    data_fim: datetime,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    if data_fim < data_inicio:
        raise HTTPException(status_code=400, detail="Data final deve ser maior ou igual a data inicial")
    if (data_fim.date() - data_inicio.date()).days > 62:
        raise HTTPException(status_code=400, detail="O relatorio geral aceita no maximo 63 dias")

    garantir_colunas_sobreaviso(db)
    subestacoes = (
        db.query(Subestacao)
        .filter(func.upper(Subestacao.status) == "ATIVA")
        .order_by(Subestacao.nome)
        .all()
    )
    colaboradores = (
        db.query(SobreavisoColaborador)
        .options(selectinload(SobreavisoColaborador.equipe))
        .filter(
            SobreavisoColaborador.ativo == 1,
            or_(
                SobreavisoColaborador.cargo.is_(None),
                func.lower(SobreavisoColaborador.cargo) != "admin",
            ),
        )
        .order_by(SobreavisoColaborador.nome)
        .all()
    )
    sobreavisos = (
        db.query(SobreavisoPeriodo)
        .filter(
            SobreavisoPeriodo.status.notin_({"REPROVADO", "CANCELADO"}),
            SobreavisoPeriodo.inicio <= data_fim,
            SobreavisoPeriodo.fim >= data_inicio,
        )
        .order_by(SobreavisoPeriodo.inicio)
        .all()
    )

    wb = montar_relatorio_escala_geral(
        subestacoes,
        colaboradores,
        sobreavisos,
        data_inicio,
        data_fim,
    )
    output_dir = Path("exports") / "sobreaviso"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"escala_geral_sobreaviso_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)

    return FileResponse(
        path=str(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.get("/{id_sobreaviso}", response_model=SobreavisoResponse)
def buscar_sobreaviso(
    id_sobreaviso: int,
    db: Session = Depends(get_db),
    _usuario=Depends(require_roles("admin", "mantenedor", "operador")),
):
    return buscar_sobreaviso_ou_404(db, id_sobreaviso)


@router.post("/", response_model=SobreavisoResponse, status_code=201)
def criar_sobreaviso(
    dados: SobreavisoCreate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    buscar_colaborador_ou_404(db, dados.id_colaborador, exigir_ativo=True)
    status = normalizar_status(dados.status)
    origem = normalizar_origem(dados.origem)

    if existe_sobreposicao(db, dados.id_colaborador, dados.inicio, dados.fim):
        raise HTTPException(status_code=400, detail="Ja existe sobreaviso no mesmo periodo para este colaborador")

    sobreaviso = SobreavisoPeriodo(
        id_colaborador=dados.id_colaborador,
        inicio=dados.inicio,
        fim=dados.fim,
        total_horas=calcular_total_horas(dados.inicio, dados.fim),
        total_horas_atendimento=Decimal("0"),
        status=status,
        origem=origem,
        justificativa=dados.justificativa,
        criado_por=usuario.id,
        atualizado_por=usuario.id,
    )
    db.add(sobreaviso)
    db.flush()
    if dados.intervalos:
        substituir_intervalos(db, sobreaviso, dados.intervalos, usuario)
    registrar_historico(
        db,
        "SOBREAVISO",
        sobreaviso.id_sobreaviso,
        "CRIACAO",
        usuario,
        dados_novos=dados.model_dump(),
        justificativa=dados.justificativa,
    )
    db.commit()
    return buscar_sobreaviso_ou_404(db, sobreaviso.id_sobreaviso)


@router.put("/{id_sobreaviso}", response_model=SobreavisoResponse)
def atualizar_sobreaviso(
    id_sobreaviso: int,
    dados: SobreavisoUpdate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    sobreaviso = buscar_sobreaviso_ou_404(db, id_sobreaviso)
    payload = dados.model_dump(exclude_unset=True)
    id_colaborador = payload.get("id_colaborador", sobreaviso.id_colaborador)
    inicio = payload.get("inicio", sobreaviso.inicio)
    fim = payload.get("fim", sobreaviso.fim)

    buscar_colaborador_ou_404(db, id_colaborador, exigir_ativo=True)
    if existe_sobreposicao(db, id_colaborador, inicio, fim, id_sobreaviso):
        raise HTTPException(status_code=400, detail="Ja existe sobreaviso no mesmo periodo para este colaborador")

    anteriores = {
        "id_colaborador": sobreaviso.id_colaborador,
        "inicio": sobreaviso.inicio,
        "fim": sobreaviso.fim,
        "status": sobreaviso.status,
        "justificativa": sobreaviso.justificativa,
    }

    intervalos = payload.pop("intervalos", None)
    for campo, valor in payload.items():
        if campo == "status" and valor is not None:
            valor = normalizar_status(valor)
        if campo == "origem" and valor is not None:
            valor = normalizar_origem(valor)
        setattr(sobreaviso, campo, valor)

    if intervalos is not None:
        substituir_intervalos(db, sobreaviso, intervalos, usuario)
    elif sobreaviso.intervalos:
        # Uma alteracao da escala nao pode deixar a linha do tempo fora da faixa.
        validar_intervalos(inicio, fim, sobreaviso.intervalos)
    else:
        sobreaviso.total_horas = calcular_total_horas(inicio, fim)
        sobreaviso.total_horas_atendimento = Decimal("0")
    sobreaviso.atualizado_por = usuario.id
    sobreaviso.atualizado_em = datetime.utcnow()

    registrar_historico(
        db,
        "SOBREAVISO",
        id_sobreaviso,
        "EDICAO",
        usuario,
        dados_anteriores=anteriores,
        dados_novos=payload,
        justificativa=sobreaviso.justificativa,
    )
    db.commit()
    return buscar_sobreaviso_ou_404(db, id_sobreaviso)


@router.delete("/{id_sobreaviso}")
def excluir_sobreaviso_cancelado(
    id_sobreaviso: int,
    db: Session = Depends(get_db),
    _usuario: Usuario = Depends(require_roles("admin")),
):
    sobreaviso = buscar_sobreaviso_ou_404(db, id_sobreaviso)
    if sobreaviso.status != "CANCELADO":
        raise HTTPException(
            status_code=400,
            detail="Somente sobreavisos cancelados podem ser excluidos",
        )

    solicitacoes_ids = [
        item[0]
        for item in db.query(SobreavisoSolicitacaoAjuste.id_solicitacao)
        .filter(SobreavisoSolicitacaoAjuste.id_sobreaviso == id_sobreaviso)
        .all()
    ]

    try:
        if solicitacoes_ids:
            db.query(SobreavisoHistorico).filter(
                SobreavisoHistorico.entidade == "SOLICITACAO_AJUSTE",
                SobreavisoHistorico.entidade_id.in_(solicitacoes_ids),
            ).delete(synchronize_session=False)

        db.query(SobreavisoSolicitacaoAjuste).filter(
            SobreavisoSolicitacaoAjuste.id_sobreaviso == id_sobreaviso
        ).delete(synchronize_session=False)
        db.query(SobreavisoHistorico).filter(
            SobreavisoHistorico.entidade == "SOBREAVISO",
            SobreavisoHistorico.entidade_id == id_sobreaviso,
        ).delete(synchronize_session=False)
        db.delete(sobreaviso)
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {"message": "Sobreaviso cancelado excluido", "id_sobreaviso": id_sobreaviso}


def alterar_status_sobreaviso(
    id_sobreaviso: int,
    status: str,
    justificativa: Optional[str],
    db: Session,
    usuario: Usuario,
):
    sobreaviso = buscar_sobreaviso_ou_404(db, id_sobreaviso)
    anterior = {"status": sobreaviso.status, "justificativa": sobreaviso.justificativa}
    sobreaviso.status = status
    if justificativa:
        sobreaviso.justificativa = justificativa
    sobreaviso.atualizado_por = usuario.id
    sobreaviso.atualizado_em = datetime.utcnow()
    registrar_historico(
        db,
        "SOBREAVISO",
        id_sobreaviso,
        status,
        usuario,
        dados_anteriores=anterior,
        dados_novos={"status": status},
        justificativa=justificativa,
    )
    db.commit()
    return buscar_sobreaviso_ou_404(db, id_sobreaviso)


@router.post("/{id_sobreaviso}/aprovar", response_model=SobreavisoResponse)
def aprovar_sobreaviso(
    id_sobreaviso: int,
    justificativa: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    return alterar_status_sobreaviso(id_sobreaviso, "APROVADO", justificativa, db, usuario)


@router.post("/{id_sobreaviso}/reprovar", response_model=SobreavisoResponse)
def reprovar_sobreaviso(
    id_sobreaviso: int,
    justificativa: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    return alterar_status_sobreaviso(id_sobreaviso, "REPROVADO", justificativa, db, usuario)


@router.post("/{id_sobreaviso}/cancelar", response_model=SobreavisoResponse)
def cancelar_sobreaviso(
    id_sobreaviso: int,
    justificativa: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    return alterar_status_sobreaviso(id_sobreaviso, "CANCELADO", justificativa, db, usuario)


@router.post("/{id_sobreaviso}/solicitar-ajuste", response_model=SolicitacaoAjusteResponse, status_code=201)
def solicitar_ajuste(
    id_sobreaviso: int,
    dados: SolicitacaoAjusteCreate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    sobreaviso = buscar_sobreaviso_ou_404(db, id_sobreaviso)
    if existe_sobreposicao(
        db,
        sobreaviso.id_colaborador,
        dados.inicio_solicitado,
        dados.fim_solicitado,
        id_sobreaviso,
    ):
        raise HTTPException(status_code=400, detail="O ajuste solicitado causaria sobreposicao de horarios")

    solicitacao = SobreavisoSolicitacaoAjuste(
        id_sobreaviso=id_sobreaviso,
        solicitado_por=usuario.id,
        inicio_solicitado=dados.inicio_solicitado,
        fim_solicitado=dados.fim_solicitado,
        justificativa=dados.justificativa,
    )
    db.add(solicitacao)
    db.flush()
    registrar_historico(
        db,
        "SOLICITACAO_AJUSTE",
        solicitacao.id_solicitacao,
        "SOLICITACAO",
        usuario,
        dados_novos=dados.model_dump(),
        justificativa=dados.justificativa,
    )
    db.commit()
    db.refresh(solicitacao)
    return solicitacao


def avaliar_solicitacao(
    id_solicitacao: int,
    status: str,
    db: Session,
    usuario: Usuario,
):
    solicitacao = (
        db.query(SobreavisoSolicitacaoAjuste)
        .filter(SobreavisoSolicitacaoAjuste.id_solicitacao == id_solicitacao)
        .first()
    )
    if not solicitacao:
        raise HTTPException(status_code=404, detail="Solicitacao de ajuste nao encontrada")
    if solicitacao.status != "PENDENTE":
        raise HTTPException(status_code=400, detail="Solicitacao ja avaliada")

    sobreaviso = buscar_sobreaviso_ou_404(db, solicitacao.id_sobreaviso)
    if status == "APROVADA":
        if existe_sobreposicao(
            db,
            sobreaviso.id_colaborador,
            solicitacao.inicio_solicitado,
            solicitacao.fim_solicitado,
            sobreaviso.id_sobreaviso,
        ):
            raise HTTPException(status_code=400, detail="A aprovacao causaria sobreposicao de horarios")
        sobreaviso.inicio = solicitacao.inicio_solicitado
        sobreaviso.fim = solicitacao.fim_solicitado
        sobreaviso.total_horas = calcular_total_horas(sobreaviso.inicio, sobreaviso.fim)
        sobreaviso.atualizado_por = usuario.id
        sobreaviso.atualizado_em = datetime.utcnow()

    solicitacao.status = status
    solicitacao.avaliado_por = usuario.id
    solicitacao.avaliado_em = datetime.utcnow()
    registrar_historico(
        db,
        "SOLICITACAO_AJUSTE",
        id_solicitacao,
        status,
        usuario,
        dados_novos={"status": status},
        justificativa=solicitacao.justificativa,
    )
    db.commit()
    db.refresh(solicitacao)
    return solicitacao


@router.post("/solicitacoes-ajuste/{id_solicitacao}/aprovar", response_model=SolicitacaoAjusteResponse)
def aprovar_solicitacao(
    id_solicitacao: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    return avaliar_solicitacao(id_solicitacao, "APROVADA", db, usuario)


@router.post("/solicitacoes-ajuste/{id_solicitacao}/reprovar", response_model=SolicitacaoAjusteResponse)
def reprovar_solicitacao(
    id_solicitacao: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_roles("admin", "mantenedor", "operador")),
):
    return avaliar_solicitacao(id_solicitacao, "REPROVADA", db, usuario)
